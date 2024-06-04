import math

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import quantstats as qs
import scipy.interpolate as sci
import seaborn as sns
import tushare as ts
from empyrical import calmar_ratio, omega_ratio, sharpe_ratio, sortino_ratio
from openpyxl import Workbook, load_workbook
from scipy.stats import ks_2samp, norm

HE_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='HE',index_col=0,parse_dates=[0])
DB_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='DB',index_col=0,parse_dates=[0])
FF_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='FF',index_col=0,parse_dates=[0])
II_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='II',index_col=0,parse_dates=[0])
ST_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='ST',index_col=0,parse_dates=[0])
CP_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='CP',index_col=0,parse_dates=[0])
FS_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='FS',index_col=0,parse_dates=[0])
FC_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='FC',index_col=0,parse_dates=[0])
AM_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='AM',index_col=0,parse_dates=[0])
SS_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\CoVaR.xlsx', sheet_name='SS',index_col=0,parse_dates=[0])

HE=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='HE',index_col=0,parse_dates=[0])
DB=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='DB',index_col=0,parse_dates=[0])
FF=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='FF',index_col=0,parse_dates=[0])
II=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='II',index_col=0,parse_dates=[0])
ST=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='ST',index_col=0,parse_dates=[0])
CP=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='CP',index_col=0,parse_dates=[0])
FS=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='FS',index_col=0,parse_dates=[0])
FC=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='FC',index_col=0,parse_dates=[0])
AM=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='AM',index_col=0,parse_dates=[0])
SS=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\CoVaR.xlsx', sheet_name='SS',index_col=0,parse_dates=[0])



#From 2 to 1-----------------------------------
NHE_Dw1=HE_N['deltaCDw1_pct'].to_numpy()
NDB_Dw1=DB_N['deltaCDw1_pct'].to_numpy()
NFF_Dw1=FF_N['deltaCDw1_pct'].to_numpy()
NII_Dw1=II_N['deltaCDw1_pct'].to_numpy()
NST_Dw1=ST_N['deltaCDw1_pct'].to_numpy()
NCP_Dw1=CP_N['deltaCDw1_pct'].to_numpy()
NFS_Dw1=FS_N['deltaCDw1_pct'].to_numpy()
NFC_Dw1=FC_N['deltaCDw1_pct'].to_numpy()
NAM_Dw1=AM_N['deltaCDw1_pct'].to_numpy()
NSS_Dw1=SS_N['deltaCDw1_pct'].to_numpy()

NHE_Up1=HE_N['deltaCUp1_pct'].to_numpy()
NDB_Up1=DB_N['deltaCUp1_pct'].to_numpy()
NFF_Up1=FF_N['deltaCUp1_pct'].to_numpy()
NII_Up1=II_N['deltaCUp1_pct'].to_numpy()
NST_Up1=ST_N['deltaCUp1_pct'].to_numpy()
NCP_Up1=CP_N['deltaCUp1_pct'].to_numpy()
NFS_Up1=FS_N['deltaCUp1_pct'].to_numpy()
NFC_Up1=FC_N['deltaCUp1_pct'].to_numpy()
NAM_Up1=AM_N['deltaCUp1_pct'].to_numpy()
NSS_Up1=SS_N['deltaCUp1_pct'].to_numpy()



HE_Dw1=HE['deltaCDw1_pct'].to_numpy()
DB_Dw1=DB['deltaCDw1_pct'].to_numpy()
FF_Dw1=FF['deltaCDw1_pct'].to_numpy()
II_Dw1=II['deltaCDw1_pct'].to_numpy()
ST_Dw1=ST['deltaCDw1_pct'].to_numpy()
CP_Dw1=CP['deltaCDw1_pct'].to_numpy()
FS_Dw1=FS['deltaCDw1_pct'].to_numpy()
FC_Dw1=FC['deltaCDw1_pct'].to_numpy()
AM_Dw1=AM['deltaCDw1_pct'].to_numpy()
SS_Dw1=SS['deltaCDw1_pct'].to_numpy()

HE_Up1=HE['deltaCUp1_pct'].to_numpy()
DB_Up1=DB['deltaCUp1_pct'].to_numpy()
FF_Up1=FF['deltaCUp1_pct'].to_numpy()
II_Up1=II['deltaCUp1_pct'].to_numpy()
ST_Up1=ST['deltaCUp1_pct'].to_numpy()
CP_Up1=CP['deltaCUp1_pct'].to_numpy()
FS_Up1=FS['deltaCUp1_pct'].to_numpy()
FC_Up1=FC['deltaCUp1_pct'].to_numpy()
AM_Up1=AM['deltaCUp1_pct'].to_numpy()
SS_Up1=SS['deltaCUp1_pct'].to_numpy()





#From 1 to 2-----------------------------------
NHE_Dw2=HE_N['deltaCDw2_pct'].to_numpy()
NDB_Dw2=DB_N['deltaCDw2_pct'].to_numpy()
NFF_Dw2=FF_N['deltaCDw2_pct'].to_numpy()
NII_Dw2=II_N['deltaCDw2_pct'].to_numpy()
NST_Dw2=ST_N['deltaCDw2_pct'].to_numpy()
NCP_Dw2=CP_N['deltaCDw2_pct'].to_numpy()
NFS_Dw2=FS_N['deltaCDw2_pct'].to_numpy()
NFC_Dw2=FC_N['deltaCDw2_pct'].to_numpy()
NAM_Dw2=AM_N['deltaCDw2_pct'].to_numpy()
NSS_Dw2=SS_N['deltaCDw2_pct'].to_numpy()

NHE_Up2=HE_N['deltaCUp2_pct'].to_numpy()
NDB_Up2=DB_N['deltaCUp2_pct'].to_numpy()
NFF_Up2=FF_N['deltaCUp2_pct'].to_numpy()
NII_Up2=II_N['deltaCUp2_pct'].to_numpy()
NST_Up2=ST_N['deltaCUp2_pct'].to_numpy()
NCP_Up2=CP_N['deltaCUp2_pct'].to_numpy()
NFS_Up2=FS_N['deltaCUp2_pct'].to_numpy()
NFC_Up2=FC_N['deltaCUp2_pct'].to_numpy()
NAM_Up2=AM_N['deltaCUp2_pct'].to_numpy()
NSS_Up2=SS_N['deltaCUp2_pct'].to_numpy()





HE_Dw2=HE['deltaCDw2_pct'].to_numpy()
DB_Dw2=DB['deltaCDw2_pct'].to_numpy()
FF_Dw2=FF['deltaCDw2_pct'].to_numpy()
II_Dw2=II['deltaCDw2_pct'].to_numpy()
ST_Dw2=ST['deltaCDw2_pct'].to_numpy()
CP_Dw2=CP['deltaCDw2_pct'].to_numpy()
FS_Dw2=FS['deltaCDw2_pct'].to_numpy()
FC_Dw2=FC['deltaCDw2_pct'].to_numpy()
AM_Dw2=AM['deltaCDw2_pct'].to_numpy()
SS_Dw2=SS['deltaCDw2_pct'].to_numpy()

HE_Up2=HE['deltaCUp2_pct'].to_numpy()
DB_Up2=DB['deltaCUp2_pct'].to_numpy()
FF_Up2=FF['deltaCUp2_pct'].to_numpy()
II_Up2=II['deltaCUp2_pct'].to_numpy()
ST_Up2=ST['deltaCUp2_pct'].to_numpy()
CP_Up2=CP['deltaCUp2_pct'].to_numpy()
FS_Up2=FS['deltaCUp2_pct'].to_numpy()
FC_Up2=FC['deltaCUp2_pct'].to_numpy()
AM_Up2=AM['deltaCUp2_pct'].to_numpy()
SS_Up2=SS['deltaCUp2_pct'].to_numpy()





# K-S-----------------------------------

wb = Workbook()


ws1 = wb.create_sheet("f2t1")
for index, value in enumerate(['NESG','ESG','S','P']*2, start=2):
    ws1.cell(row=2, column=index, value=value)
ws1['B1']='Downside'
ws1['F1']='Upside'
ws1['A3']='HE'
ws1['A4']='DB'
ws1['A5']='FF'
ws1['A6']='II'
ws1['A7']='ST'
ws1['A8']='CP'
ws1['A9']='FS'
ws1['A10']='FC'
ws1['A11']='AM'
ws1['A12']='SS'
for index, value in enumerate([NHE_Dw1.mean(),HE_Dw1.mean(),ks_2samp(NHE_Dw1,HE_Dw1)[0],ks_2samp(NHE_Dw1,HE_Dw1)[1],
                               NHE_Up1.mean(),HE_Up1.mean(),ks_2samp(NHE_Up1,HE_Up1)[0],ks_2samp(NHE_Up1,HE_Up1)[1]], start=2):
    ws1.cell(row=3, column=index, value=value)

for index, value in enumerate([NDB_Dw1.mean(),DB_Dw1.mean(),ks_2samp(NDB_Dw1,DB_Dw1)[0],ks_2samp(NDB_Dw1,DB_Dw1)[1],
                               NDB_Up1.mean(),DB_Up1.mean(),ks_2samp(NDB_Up1,DB_Up1)[0],ks_2samp(NDB_Up1,DB_Up1)[1]], start=2):
    ws1.cell(row=4, column=index, value=value)

for index, value in enumerate([NFF_Dw1.mean(),FF_Dw1.mean(),ks_2samp(NFF_Dw1,FF_Dw1)[0],ks_2samp(NFF_Dw1,FF_Dw1)[1],
                               NFF_Up1.mean(),FF_Up1.mean(),ks_2samp(NFF_Up1,FF_Up1)[0],ks_2samp(NFF_Up1,FF_Up1)[1]], start=2):
    ws1.cell(row=5, column=index, value=value)

for index, value in enumerate([NII_Dw1.mean(),II_Dw1.mean(),ks_2samp(NII_Dw1,II_Dw1)[0],ks_2samp(NII_Dw1,II_Dw1)[1],
                               NII_Up1.mean(),II_Up1.mean(),ks_2samp(NII_Up1,II_Up1)[0],ks_2samp(NII_Up1,II_Up1)[1]], start=2):
    ws1.cell(row=6, column=index, value=value)

for index, value in enumerate([NST_Dw1.mean(),ST_Dw1.mean(),ks_2samp(NST_Dw1,ST_Dw1)[0],ks_2samp(NST_Dw1,ST_Dw1)[1],
                               NST_Up1.mean(),ST_Up1.mean(),ks_2samp(NST_Up1,ST_Up1)[0],ks_2samp(NST_Up1,ST_Up1)[1]], start=2):
    ws1.cell(row=7, column=index, value=value)

for index, value in enumerate([NCP_Dw1.mean(),CP_Dw1.mean(),ks_2samp(NCP_Dw1,CP_Dw1)[0],ks_2samp(NCP_Dw1,CP_Dw1)[1],
                               NCP_Up1.mean(),CP_Up1.mean(),ks_2samp(NCP_Up1,CP_Up1)[0],ks_2samp(NCP_Up1,CP_Up1)[1]], start=2):
    ws1.cell(row=8, column=index, value=value)

for index, value in enumerate([NFS_Dw1.mean(),FS_Dw1.mean(),ks_2samp(NFS_Dw1,FS_Dw1)[0],ks_2samp(NFS_Dw1,FS_Dw1)[1],
                               NFS_Up1.mean(),FS_Up1.mean(),ks_2samp(NFS_Up1,FS_Up1)[0],ks_2samp(NFS_Up1,FS_Up1)[1]], start=2):
    ws1.cell(row=9, column=index, value=value)

for index, value in enumerate([NFC_Dw1.mean(),FC_Dw1.mean(),ks_2samp(NFC_Dw1,FC_Dw1)[0],ks_2samp(NFC_Dw1,FC_Dw1)[1],
                               NFC_Up1.mean(),FC_Up1.mean(),ks_2samp(NFC_Up1,FC_Up1)[0],ks_2samp(NFC_Up1,FC_Up1)[1]], start=2):
    ws1.cell(row=10, column=index, value=value)

for index, value in enumerate([NAM_Dw1.mean(),AM_Dw1.mean(),ks_2samp(NAM_Dw1,AM_Dw1)[0],ks_2samp(NAM_Dw1,AM_Dw1)[1],
                               NAM_Up1.mean(),AM_Up1.mean(),ks_2samp(NAM_Up1,AM_Up1)[0],ks_2samp(NAM_Up1,AM_Up1)[1]], start=2):
    ws1.cell(row=11, column=index, value=value)

for index, value in enumerate([NSS_Dw1.mean(),SS_Dw1.mean(),ks_2samp(NSS_Dw1,SS_Dw1)[0],ks_2samp(NSS_Dw1,SS_Dw1)[1],
                               NSS_Up1.mean(),SS_Up1.mean(),ks_2samp(NSS_Up1,SS_Up1)[0],ks_2samp(NSS_Up1,SS_Up1)[1]], start=2):
    ws1.cell(row=12, column=index, value=value)







ws2 = wb.create_sheet("f1t2")
for index, value in enumerate(['NESG','ESG','S','P']*2, start=2):
    ws2.cell(row=2, column=index, value=value)
ws2['B1']='Downside'
ws2['F1']='Upside'
ws2['A3']='HE'
ws2['A4']='DB'
ws2['A5']='FF'
ws2['A6']='II'
ws2['A7']='ST'
ws2['A8']='CP'
ws2['A9']='FS'
ws2['A10']='FC'
ws2['A11']='AM'
ws2['A12']='SS'
for index, value in enumerate([NHE_Dw2.mean(),HE_Dw2.mean(),ks_2samp(NHE_Dw2,HE_Dw2)[0],ks_2samp(NHE_Dw2,HE_Dw2)[1],
                               NHE_Up2.mean(),HE_Up2.mean(),ks_2samp(NHE_Up2,HE_Up2)[0],ks_2samp(NHE_Up2,HE_Up2)[1]], start=2):
    ws2.cell(row=3, column=index, value=value)

for index, value in enumerate([NDB_Dw2.mean(),DB_Dw2.mean(),ks_2samp(NDB_Dw2,DB_Dw2)[0],ks_2samp(NDB_Dw2,DB_Dw2)[1],
                               NDB_Up2.mean(),DB_Up2.mean(),ks_2samp(NDB_Up2,DB_Up2)[0],ks_2samp(NDB_Up2,DB_Up2)[1]], start=2):
    ws2.cell(row=4, column=index, value=value)

for index, value in enumerate([NFF_Dw2.mean(),FF_Dw2.mean(),ks_2samp(NFF_Dw2,FF_Dw2)[0],ks_2samp(NFF_Dw2,FF_Dw2)[1],
                               NFF_Up2.mean(),FF_Up2.mean(),ks_2samp(NFF_Up2,FF_Up2)[0],ks_2samp(NFF_Up2,FF_Up2)[1]], start=2):
    ws2.cell(row=5, column=index, value=value)

for index, value in enumerate([NII_Dw2.mean(),II_Dw2.mean(),ks_2samp(NII_Dw2,II_Dw2)[0],ks_2samp(NII_Dw2,II_Dw2)[1],
                               NII_Up2.mean(),II_Up2.mean(),ks_2samp(NII_Up2,II_Up2)[0],ks_2samp(NII_Up2,II_Up2)[1]], start=2):
    ws2.cell(row=6, column=index, value=value)

for index, value in enumerate([NST_Dw2.mean(),ST_Dw2.mean(),ks_2samp(NST_Dw2,ST_Dw2)[0],ks_2samp(NST_Dw2,ST_Dw2)[1],
                               NST_Up2.mean(),ST_Up2.mean(),ks_2samp(NST_Up2,ST_Up2)[0],ks_2samp(NST_Up2,ST_Up2)[1]], start=2):
    ws2.cell(row=7, column=index, value=value)

for index, value in enumerate([NCP_Dw2.mean(),CP_Dw2.mean(),ks_2samp(NCP_Dw2,CP_Dw2)[0],ks_2samp(NCP_Dw2,CP_Dw2)[1],
                               NCP_Up2.mean(),CP_Up2.mean(),ks_2samp(NCP_Up2,CP_Up2)[0],ks_2samp(NCP_Up2,CP_Up2)[1]], start=2):
    ws2.cell(row=8, column=index, value=value)

for index, value in enumerate([NFS_Dw2.mean(),FS_Dw2.mean(),ks_2samp(NFS_Dw2,FS_Dw2)[0],ks_2samp(NFS_Dw2,FS_Dw2)[1],
                               NFS_Up2.mean(),FS_Up2.mean(),ks_2samp(NFS_Up2,FS_Up2)[0],ks_2samp(NFS_Up2,FS_Up2)[1]], start=2):
    ws2.cell(row=9, column=index, value=value)

for index, value in enumerate([NFC_Dw2.mean(),FC_Dw2.mean(),ks_2samp(NFC_Dw2,FC_Dw2)[0],ks_2samp(NFC_Dw2,FC_Dw2)[1],
                               NFC_Up2.mean(),FC_Up2.mean(),ks_2samp(NFC_Up2,FC_Up2)[0],ks_2samp(NFC_Up2,FC_Up2)[1]], start=2):
    ws2.cell(row=10, column=index, value=value)

for index, value in enumerate([NAM_Dw2.mean(),AM_Dw2.mean(),ks_2samp(NAM_Dw2,AM_Dw2)[0],ks_2samp(NAM_Dw2,AM_Dw2)[1],
                               NAM_Up2.mean(),AM_Up2.mean(),ks_2samp(NAM_Up2,AM_Up2)[0],ks_2samp(NAM_Up2,AM_Up2)[1]], start=2):
    ws2.cell(row=11, column=index, value=value)

for index, value in enumerate([NSS_Dw2.mean(),SS_Dw2.mean(),ks_2samp(NSS_Dw2,SS_Dw2)[0],ks_2samp(NSS_Dw2,SS_Dw2)[1],
                               NSS_Up2.mean(),SS_Up2.mean(),ks_2samp(NSS_Up2,SS_Up2)[0],ks_2samp(NSS_Up2,SS_Up2)[1]], start=2):
    ws2.cell(row=12, column=index, value=value)


wb.save(r'C:\Users\徐浩然\Desktop\总\KS.xlsx')