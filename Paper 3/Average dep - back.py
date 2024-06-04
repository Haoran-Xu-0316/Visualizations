import math

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import quantstats as qs
import scipy.interpolate as sci
import seaborn as sns
import tushare as ts
from empyrical import calmar_ratio, omega_ratio, sharpe_ratio, sortino_ratio
from openpyxl import Workbook, load_workbook
from scipy.stats import ks_2samp, norm

HE_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='HE',index_col=0,parse_dates=[0])
DB_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='DB',index_col=0,parse_dates=[0])
FF_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='FF',index_col=0,parse_dates=[0])
II_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='II',index_col=0,parse_dates=[0])
ST_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='ST',index_col=0,parse_dates=[0])
CP_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='CP',index_col=0,parse_dates=[0])
FS_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='FS',index_col=0,parse_dates=[0])
FC_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='FC',index_col=0,parse_dates=[0])
AM_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='AM',index_col=0,parse_dates=[0])
SS_N=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 NOESG\Tau_Taildep.xlsx', sheet_name='SS',index_col=0,parse_dates=[0])

HE=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='HE',index_col=0,parse_dates=[0])
DB=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='DB',index_col=0,parse_dates=[0])
FF=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='FF',index_col=0,parse_dates=[0])
II=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='II',index_col=0,parse_dates=[0])
ST=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='ST',index_col=0,parse_dates=[0])
CP=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='CP',index_col=0,parse_dates=[0])
FS=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='FS',index_col=0,parse_dates=[0])
FC=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='FC',index_col=0,parse_dates=[0])
AM=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='AM',index_col=0,parse_dates=[0])
SS=pd.read_excel(r'C:\Users\徐浩然\Desktop\P4 ESG\Tau_Taildep.xlsx', sheet_name='SS',index_col=0,parse_dates=[0])



#From 2 to 1-----------------------------------
NHE_1=HE_N['tau'].to_numpy()
NDB_1=DB_N['tau'].to_numpy()
NFF_1=FF_N['tau'].to_numpy()
NII_1=II_N['tau'].to_numpy()
NST_1=ST_N['tau'].to_numpy()
NCP_1=CP_N['tau'].to_numpy()
NFS_1=FS_N['tau'].to_numpy()
NFC_1=FC_N['tau'].to_numpy()
NAM_1=AM_N['tau'].to_numpy()
NSS_1=SS_N['tau'].to_numpy()

NHE_2=HE_N['uptail'].to_numpy()
NDB_2=DB_N['uptail'].to_numpy()
NFF_2=FF_N['uptail'].to_numpy()
NII_2=II_N['uptail'].to_numpy()
NST_2=ST_N['uptail'].to_numpy()
NCP_2=CP_N['uptail'].to_numpy()
NFS_2=FS_N['uptail'].to_numpy()
NFC_2=FC_N['uptail'].to_numpy()
NAM_2=AM_N['uptail'].to_numpy()
NSS_2=SS_N['uptail'].to_numpy()



HE_1=HE['tau'].to_numpy()
DB_1=DB['tau'].to_numpy()
FF_1=FF['tau'].to_numpy()
II_1=II['tau'].to_numpy()
ST_1=ST['tau'].to_numpy()
CP_1=CP['tau'].to_numpy()
FS_1=FS['tau'].to_numpy()
FC_1=FC['tau'].to_numpy()
AM_1=AM['tau'].to_numpy()
SS_1=SS['tau'].to_numpy()

HE_2=HE['uptail'].to_numpy()
DB_2=DB['uptail'].to_numpy()
FF_2=FF['uptail'].to_numpy()
II_2=II['uptail'].to_numpy()
ST_2=ST['uptail'].to_numpy()
CP_2=CP['uptail'].to_numpy()
FS_2=FS['uptail'].to_numpy()
FC_2=FC['uptail'].to_numpy()
AM_2=AM['uptail'].to_numpy()
SS_2=SS['uptail'].to_numpy()




# K-S-----------------------------------

wb = Workbook()


ws = wb.create_sheet("dep平均")
for index, value in enumerate(['tau','taildep']*2, start=2):
    ws.cell(row=2, column=index, value=value)
ws['B1']='NESG'
ws['D1']='ESG'
ws['A3']='HE'
ws['A4']='DB'
ws['A5']='FF'
ws['A6']='II'
ws['A7']='ST'
ws['A8']='CP'
ws['A9']='FS'
ws['A10']='FC'
ws['A11']='AM'
ws['A12']='SS'



for index, value in enumerate([NHE_1.mean(), NHE_2.mean(), HE_1.mean(), HE_2.mean()], start=2):
    ws.cell(row=3, column=index, value=value)

for index, value in enumerate([NDB_1.mean(), NDB_2.mean(), DB_1.mean(), DB_2.mean()], start=2):
    ws.cell(row=5, column=index, value=value)

for index, value in enumerate([NFF_1.mean(), NFF_2.mean(), FF_1.mean(), FF_2.mean()], start=2):
    ws.cell(row=7, column=index, value=value)

for index, value in enumerate([NII_1.mean(), NII_2.mean(), II_1.mean(), II_2.mean()], start=2):
    ws.cell(row=9, column=index, value=value)

for index, value in enumerate([NST_1.mean(), NST_2.mean(), ST_1.mean(), ST_2.mean()], start=2):
    ws.cell(row=11, column=index, value=value)

for index, value in enumerate([NCP_1.mean(), NCP_2.mean(), CP_1.mean(), CP_2.mean()], start=2):
    ws.cell(row=13, column=index, value=value)

for index, value in enumerate([NFS_1.mean(), NFS_2.mean(), FS_1.mean(), FS_2.mean()], start=2):
    ws.cell(row=15, column=index, value=value)

for index, value in enumerate([NFC_1.mean(), NFC_2.mean(), FC_1.mean(), FC_2.mean()], start=2):
    ws.cell(row=17, column=index, value=value)

for index, value in enumerate([NAM_1.mean(), NAM_2.mean(), AM_1.mean(), AM_2.mean()], start=2):
    ws.cell(row=19, column=index, value=value)

for index, value in enumerate([NSS_1.mean(), NSS_2.mean(), SS_1.mean(), SS_2.mean()], start=2):
    ws.cell(row=21, column=index, value=value)








for index, value in enumerate([NHE_1.std(), NHE_2.std(), HE_1.std(), HE_2.std()], start=2):
    ws.cell(row=4, column=index, value=value)

for index, value in enumerate([NDB_1.std(), NDB_2.std(), DB_1.std(), DB_2.std()], start=2):
    ws.cell(row=6, column=index, value=value)

for index, value in enumerate([NFF_1.std(), NFF_2.std(), FF_1.std(), FF_2.std()], start=2):
    ws.cell(row=8, column=index, value=value)

for index, value in enumerate([NII_1.std(), NII_2.std(), II_1.std(), II_2.std()], start=2):
    ws.cell(row=10, column=index, value=value)

for index, value in enumerate([NST_1.std(), NST_2.std(), ST_1.std(), ST_2.std()], start=2):
    ws.cell(row=12, column=index, value=value)

for index, value in enumerate([NCP_1.std(), NCP_2.std(), CP_1.std(), CP_2.std()], start=2):
    ws.cell(row=14, column=index, value=value)

for index, value in enumerate([NFS_1.std(), NFS_2.std(), FS_1.std(), FS_2.std()], start=2):
    ws.cell(row=16, column=index, value=value)

for index, value in enumerate([NFC_1.std(), NFC_2.std(), FC_1.std(), FC_2.std()], start=2):
    ws.cell(row=18, column=index, value=value)

for index, value in enumerate([NAM_1.std(), NAM_2.std(), AM_1.std(), AM_2.std()], start=2):
    ws.cell(row=20, column=index, value=value)

for index, value in enumerate([NSS_1.std(), NSS_2.std(), SS_1.std(), SS_2.std()], start=2):
    ws.cell(row=22, column=index, value=value)





wb.save(r'C:\Users\徐浩然\Desktop\总\dep mean_back.xlsx')