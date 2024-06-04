import math
from datetime import datetime

import matplotlib.dates as dates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import quantstats as qs
import scipy.interpolate as sci
import scipy.optimize as sco
import seaborn as sns
import tushare as ts
from empyrical import calmar_ratio, omega_ratio, sharpe_ratio, sortino_ratio
from openpyxl import Workbook, load_workbook
from scipy.stats import norm

# 全局设置-----------------------------------------------------------------------------------------------------------------
# plt.rcParams['font.sans-serif']=['Microsoft YaHei']
plt.rcParams['font.sans-serif'] = ['Palatino Linotype']
# plt.rcParams["font.weight"] = "bold"
plt.rcParams['font.size'] = 12
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['lines.linewidth'] = 1
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

HA = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedging Assets\hedging assets.xlsx', sheet_name='All-log',
                   index_col=0, parse_dates=[0])
StockCH = HA['CSI 300']
StockUS = HA['S&P 500']

Oil = HA['WTI']
BTC = HA['Bitcoin']
Gold = HA['Gold']
Bond = HA['Bond']

S_CPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-CPU.xlsx', sheet_name='Short-Cor',
                      index_col=0, parse_dates=[0])
S_EPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EPU.xlsx', sheet_name='Short-Cor',
                      index_col=0, parse_dates=[0])
S_TPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-TPU.xlsx', sheet_name='Short-Cor',
                      index_col=0, parse_dates=[0])
S_EMV = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EMV.xlsx', sheet_name='Short-Cor',
                      index_col=0, parse_dates=[0])
S_GPR = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-GPR.xlsx', sheet_name='Short-Cor',
                      index_col=0, parse_dates=[0])
S_bs = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-Baseline.xlsx',
                     sheet_name='Short-Cor', index_col=0, parse_dates=[0])

L_CPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-CPU.xlsx', sheet_name='Long-Cor',
                      index_col=0, parse_dates=[0])
L_EPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EPU.xlsx', sheet_name='Long-Cor',
                      index_col=0, parse_dates=[0])
L_TPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-TPU.xlsx', sheet_name='Long-Cor',
                      index_col=0, parse_dates=[0])
L_EMV = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EMV.xlsx', sheet_name='Long-Cor',
                      index_col=0, parse_dates=[0])
L_GPR = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-GPR.xlsx', sheet_name='Long-Cor',
                      index_col=0, parse_dates=[0])

S_bond_CPU = S_CPU.iloc[:, 3]
S_gold_CPU = S_CPU.iloc[:, 0]
S_oil_CPU = S_CPU.iloc[:, 1]
S_btc_CPU = S_CPU.iloc[:, 2]

S_bond_EPU = S_EPU.iloc[:, 3]
S_gold_EPU = S_EPU.iloc[:, 0]
S_oil_EPU = S_EPU.iloc[:, 1]
S_btc_EPU = S_EPU.iloc[:, 2]

S_bond_TPU = S_TPU.iloc[:, 3]
S_gold_TPU = S_TPU.iloc[:, 0]
S_oil_TPU = S_TPU.iloc[:, 1]
S_btc_TPU = S_TPU.iloc[:, 2]

S_bond_GPR = S_GPR.iloc[:, 3]
S_gold_GPR = S_GPR.iloc[:, 0]
S_oil_GPR = S_GPR.iloc[:, 1]
S_btc_GPR = S_GPR.iloc[:, 2]

S_bond_EMV = S_EMV.iloc[:, 3]
S_gold_EMV = S_EMV.iloc[:, 0]
S_oil_EMV = S_EMV.iloc[:, 1]
S_btc_EMV = S_EMV.iloc[:, 2]

# bs
S_bond_bs = S_bs.iloc[:, 3]
S_gold_bs = S_bs.iloc[:, 0]
S_oil_bs = S_bs.iloc[:, 1]
S_btc_bs = S_bs.iloc[:, 2]

L_bond_CPU = L_CPU.iloc[:, 3]
L_gold_CPU = L_CPU.iloc[:, 0]
L_oil_CPU = L_CPU.iloc[:, 1]
L_btc_CPU = L_CPU.iloc[:, 2]

L_bond_EPU = L_EPU.iloc[:, 3]
L_gold_EPU = L_EPU.iloc[:, 0]
L_oil_EPU = L_EPU.iloc[:, 1]
L_btc_EPU = L_EPU.iloc[:, 2]

L_bond_TPU = L_TPU.iloc[:, 3]
L_gold_TPU = L_TPU.iloc[:, 0]
L_oil_TPU = L_TPU.iloc[:, 1]
L_btc_TPU = L_TPU.iloc[:, 2]

L_bond_GPR = L_GPR.iloc[:, 3]
L_gold_GPR = L_GPR.iloc[:, 0]
L_oil_GPR = L_GPR.iloc[:, 1]
L_btc_GPR = L_GPR.iloc[:, 2]

L_bond_EMV = L_EMV.iloc[:, 3]
L_gold_EMV = L_EMV.iloc[:, 0]
L_oil_EMV = L_EMV.iloc[:, 1]
L_btc_EMV = L_EMV.iloc[:, 2]

Var_CPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-CPU.xlsx', sheet_name='Var',
                        index_col=0, parse_dates=[0])
Var_EMV = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EMV.xlsx', sheet_name='Var',
                        index_col=0, parse_dates=[0])
Var_EPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EPU.xlsx', sheet_name='Var',
                        index_col=0, parse_dates=[0])
Var_TPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-TPU.xlsx', sheet_name='Var',
                        index_col=0, parse_dates=[0])
Var_GPR = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-GPR.xlsx', sheet_name='Var',
                        index_col=0, parse_dates=[0])
Var_bs = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-Baseline.xlsx', sheet_name='Var',
                       index_col=0, parse_dates=[0])

CoV_CPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-CPU.xlsx', sheet_name='Covar',
                        index_col=0, parse_dates=[0])
CoV_EMV = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EMV.xlsx', sheet_name='Covar',
                        index_col=0, parse_dates=[0])
CoV_EPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-EPU.xlsx', sheet_name='Covar',
                        index_col=0, parse_dates=[0])
CoV_TPU = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-TPU.xlsx', sheet_name='Covar',
                        index_col=0, parse_dates=[0])
CoV_GPR = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-GPR.xlsx', sheet_name='Covar',
                        index_col=0, parse_dates=[0])
CoV_bs = pd.read_excel(r'C:\Users\徐浩然\Desktop\DCC-data-China\Varanice&Covariance-Baseline.xlsx', sheet_name='Covar',
                       index_col=0, parse_dates=[0])

## Var & CoV----CPU------------------------------------------------------------------------------------------------------------
Var_CPU_CSI300 = Var_CPU.iloc[:, 0]
Var_CPU_Gold = Var_CPU.iloc[:, 1]
Var_CPU_Oil = Var_CPU.iloc[:, 2]
Var_CPU_BTC = Var_CPU.iloc[:, 3]
Var_CPU_Bond = Var_CPU.iloc[:, 4]

CoV_CPU_Gold = CoV_CPU.iloc[:, 0]
CoV_CPU_Oil = CoV_CPU.iloc[:, 1]
CoV_CPU_BTC = CoV_CPU.iloc[:, 2]
CoV_CPU_Bond = CoV_CPU.iloc[:, 3]

## Var & CoV----EMV------------------------------------------------------------------------------------------------------------
Var_EMV_CSI300 = Var_EMV.iloc[:, 0]
Var_EMV_Gold = Var_EMV.iloc[:, 1]
Var_EMV_Oil = Var_EMV.iloc[:, 2]
Var_EMV_BTC = Var_EMV.iloc[:, 3]
Var_EMV_Bond = Var_EMV.iloc[:, 4]

CoV_EMV_Gold = CoV_EMV.iloc[:, 0]
CoV_EMV_Oil = CoV_EMV.iloc[:, 1]
CoV_EMV_BTC = CoV_EMV.iloc[:, 2]
CoV_EMV_Bond = CoV_EMV.iloc[:, 3]

## Var & CoV----EPU------------------------------------------------------------------------------------------------------------
Var_EPU_CSI300 = Var_EPU.iloc[:, 0]
Var_EPU_Gold = Var_EPU.iloc[:, 1]
Var_EPU_Oil = Var_EPU.iloc[:, 2]
Var_EPU_BTC = Var_EPU.iloc[:, 3]
Var_EPU_Bond = Var_EPU.iloc[:, 4]

CoV_EPU_Gold = CoV_EPU.iloc[:, 0]
CoV_EPU_Oil = CoV_EPU.iloc[:, 1]
CoV_EPU_BTC = CoV_EPU.iloc[:, 2]
CoV_EPU_Bond = CoV_EPU.iloc[:, 3]

## Var & CoV----TPU------------------------------------------------------------------------------------------------------------
Var_TPU_CSI300 = Var_TPU.iloc[:, 0]
Var_TPU_Gold = Var_TPU.iloc[:, 1]
Var_TPU_Oil = Var_TPU.iloc[:, 2]
Var_TPU_BTC = Var_TPU.iloc[:, 3]
Var_TPU_Bond = Var_TPU.iloc[:, 4]

CoV_TPU_Gold = CoV_TPU.iloc[:, 0]
CoV_TPU_Oil = CoV_TPU.iloc[:, 1]
CoV_TPU_BTC = CoV_TPU.iloc[:, 2]
CoV_TPU_Bond = CoV_TPU.iloc[:, 3]

## Var & CoV----GPR------------------------------------------------------------------------------------------------------------
Var_GPR_CSI300 = Var_GPR.iloc[:, 0]
Var_GPR_Gold = Var_GPR.iloc[:, 1]
Var_GPR_Oil = Var_GPR.iloc[:, 2]
Var_GPR_BTC = Var_GPR.iloc[:, 3]
Var_GPR_Bond = Var_GPR.iloc[:, 4]

CoV_GPR_Gold = CoV_GPR.iloc[:, 0]
CoV_GPR_Oil = CoV_GPR.iloc[:, 1]
CoV_GPR_BTC = CoV_GPR.iloc[:, 2]
CoV_GPR_Bond = CoV_GPR.iloc[:, 3]

## Var & CoV----base------------------------------------------------------------------------------------------------------------
Var_bs_CSI300 = Var_bs.iloc[:, 0]
Var_bs_Gold = Var_bs.iloc[:, 1]
Var_bs_Oil = Var_bs.iloc[:, 2]
Var_bs_BTC = Var_bs.iloc[:, 3]
Var_bs_Bond = Var_bs.iloc[:, 4]

CoV_bs_Gold = CoV_bs.iloc[:, 0]
CoV_bs_Oil = CoV_bs.iloc[:, 1]
CoV_bs_BTC = CoV_bs.iloc[:, 2]
CoV_bs_Bond = CoV_bs.iloc[:, 3]

## CPU OW---------------------------------------------------------------------------------------------------------------
Hedge_gold_CPU_OW = (Var_CPU_CSI300 - CoV_CPU_Gold) / (Var_CPU_CSI300 + Var_CPU_Gold - 2 * CoV_CPU_Gold)
Hedge_oil_CPU_OW = (Var_CPU_CSI300 - CoV_CPU_Oil) / (Var_CPU_CSI300 + Var_CPU_Oil - 2 * CoV_CPU_Oil)
Hedge_btc_CPU_OW = (Var_CPU_CSI300 - CoV_CPU_BTC) / (Var_CPU_CSI300 + Var_CPU_BTC - 2 * CoV_CPU_BTC)
Hedge_bond_CPU_OW = (Var_CPU_CSI300 - CoV_CPU_Bond) / (Var_CPU_CSI300 + Var_CPU_Bond - 2 * CoV_CPU_Bond)

Hedge_gold_CPU_OW = np.clip(Hedge_gold_CPU_OW, 0, 1)
Hedge_oil_CPU_OW = np.clip(Hedge_oil_CPU_OW, 0, 1)
Hedge_btc_CPU_OW = np.clip(Hedge_btc_CPU_OW, 0, 1)
Hedge_bond_CPU_OW = np.clip(Hedge_bond_CPU_OW, 0, 1)

## EMV OW---------------------------------------------------------------------------------------------------------------
Hedge_gold_EMV_OW = (Var_EMV_CSI300 - CoV_EMV_Gold) / (Var_EMV_CSI300 + Var_EMV_Gold - 2 * CoV_EMV_Gold)
Hedge_oil_EMV_OW = (Var_EMV_CSI300 - CoV_EMV_Oil) / (Var_EMV_CSI300 + Var_EMV_Oil - 2 * CoV_EMV_Oil)
Hedge_btc_EMV_OW = (Var_EMV_CSI300 - CoV_EMV_BTC) / (Var_EMV_CSI300 + Var_EMV_BTC - 2 * CoV_EMV_BTC)
Hedge_bond_EMV_OW = (Var_EMV_CSI300 - CoV_EMV_Bond) / (Var_EMV_CSI300 + Var_EMV_Bond - 2 * CoV_EMV_Bond)

Hedge_gold_EMV_OW = np.clip(Hedge_gold_EMV_OW, 0, 1)
Hedge_oil_EMV_OW = np.clip(Hedge_oil_EMV_OW, 0, 1)
Hedge_btc_EMV_OW = np.clip(Hedge_btc_EMV_OW, 0, 1)
Hedge_bond_EMV_OW = np.clip(Hedge_bond_EMV_OW, 0, 1)
## EPU OW---------------------------------------------------------------------------------------------------------------
Hedge_gold_EPU_OW = (Var_EPU_CSI300 - CoV_EPU_Gold) / (Var_EPU_CSI300 + Var_EPU_Gold - 2 * CoV_EPU_Gold)
Hedge_oil_EPU_OW = (Var_EPU_CSI300 - CoV_EPU_Oil) / (Var_EPU_CSI300 + Var_EPU_Oil - 2 * CoV_EPU_Oil)
Hedge_btc_EPU_OW = (Var_EPU_CSI300 - CoV_EPU_BTC) / (Var_EPU_CSI300 + Var_EPU_BTC - 2 * CoV_EPU_BTC)
Hedge_bond_EPU_OW = (Var_EPU_CSI300 - CoV_EPU_Bond) / (Var_EPU_CSI300 + Var_EPU_Bond - 2 * CoV_EPU_Bond)

Hedge_gold_EPU_OW = np.clip(Hedge_gold_EPU_OW, 0, 1)
Hedge_oil_EPU_OW = np.clip(Hedge_oil_EPU_OW, 0, 1)
Hedge_btc_EPU_OW = np.clip(Hedge_btc_EPU_OW, 0, 1)
Hedge_bond_EPU_OW = np.clip(Hedge_bond_EPU_OW, 0, 1)

## TPU OW---------------------------------------------------------------------------------------------------------------
Hedge_gold_TPU_OW = (Var_TPU_CSI300 - CoV_TPU_Gold) / (Var_TPU_CSI300 + Var_TPU_Gold - 2 * CoV_TPU_Gold)
Hedge_oil_TPU_OW = (Var_TPU_CSI300 - CoV_TPU_Oil) / (Var_TPU_CSI300 + Var_TPU_Oil - 2 * CoV_TPU_Oil)
Hedge_btc_TPU_OW = (Var_TPU_CSI300 - CoV_TPU_BTC) / (Var_TPU_CSI300 + Var_TPU_BTC - 2 * CoV_TPU_BTC)
Hedge_bond_TPU_OW = (Var_TPU_CSI300 - CoV_TPU_Bond) / (Var_TPU_CSI300 + Var_TPU_Bond - 2 * CoV_TPU_Bond)

Hedge_gold_TPU_OW = np.clip(Hedge_gold_TPU_OW, 0, 1)
Hedge_oil_TPU_OW = np.clip(Hedge_oil_TPU_OW, 0, 1)
Hedge_btc_TPU_OW = np.clip(Hedge_btc_TPU_OW, 0, 1)
Hedge_bond_TPU_OW = np.clip(Hedge_bond_TPU_OW, 0, 1)
## GPR OW---------------------------------------------------------------------------------------------------------------
Hedge_gold_GPR_OW = (Var_GPR_CSI300 - CoV_GPR_Gold) / (Var_GPR_CSI300 + Var_GPR_Gold - 2 * CoV_GPR_Gold)
Hedge_oil_GPR_OW = (Var_GPR_CSI300 - CoV_GPR_Oil) / (Var_GPR_CSI300 + Var_GPR_Oil - 2 * CoV_GPR_Oil)
Hedge_btc_GPR_OW = (Var_GPR_CSI300 - CoV_GPR_BTC) / (Var_GPR_CSI300 + Var_GPR_BTC - 2 * CoV_GPR_BTC)
Hedge_bond_GPR_OW = (Var_GPR_CSI300 - CoV_GPR_Bond) / (Var_GPR_CSI300 + Var_GPR_Bond - 2 * CoV_GPR_Bond)

Hedge_gold_GPR_OW = np.clip(Hedge_gold_GPR_OW, 0, 1)
Hedge_oil_GPR_OW = np.clip(Hedge_oil_GPR_OW, 0, 1)
Hedge_btc_GPR_OW = np.clip(Hedge_btc_GPR_OW, 0, 1)
Hedge_bond_GPR_OW = np.clip(Hedge_bond_GPR_OW, 0, 1)

## BS OW---------------------------------------------------------------------------------------------------------------
Hedge_gold_bs_OW = (Var_bs_CSI300 - CoV_bs_Gold) / (Var_bs_CSI300 + Var_bs_Gold - 2 * CoV_bs_Gold)
Hedge_oil_bs_OW = (Var_bs_CSI300 - CoV_bs_Oil) / (Var_bs_CSI300 + Var_bs_Oil - 2 * CoV_bs_Oil)
Hedge_btc_bs_OW = (Var_bs_CSI300 - CoV_bs_BTC) / (Var_bs_CSI300 + Var_bs_BTC - 2 * CoV_bs_BTC)
Hedge_bond_bs_OW = (Var_bs_CSI300 - CoV_bs_Bond) / (Var_bs_CSI300 + Var_bs_Bond - 2 * CoV_bs_Bond)

Hedge_gold_bs_OW = np.clip(Hedge_gold_bs_OW, 0, 1)
Hedge_oil_bs_OW = np.clip(Hedge_oil_bs_OW, 0, 1)
Hedge_btc_bs_OW = np.clip(Hedge_btc_bs_OW, 0, 1)
Hedge_bond_bs_OW = np.clip(Hedge_bond_bs_OW, 0, 1)

OW_Gold = pd.concat(
    [Hedge_gold_TPU_OW, Hedge_gold_EMV_OW, Hedge_gold_EPU_OW, Hedge_gold_GPR_OW, Hedge_gold_CPU_OW, Hedge_gold_bs_OW],
    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
OW_Oil = pd.concat(
    [Hedge_oil_TPU_OW, Hedge_oil_EMV_OW, Hedge_oil_EPU_OW, Hedge_oil_GPR_OW, Hedge_oil_CPU_OW, Hedge_oil_bs_OW],
    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
OW_BTC = pd.concat(
    [Hedge_btc_TPU_OW, Hedge_btc_EMV_OW, Hedge_btc_EPU_OW, Hedge_btc_GPR_OW, Hedge_btc_CPU_OW, Hedge_btc_bs_OW],
    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
OW_Bond = pd.concat(
    [Hedge_bond_TPU_OW, Hedge_bond_EMV_OW, Hedge_bond_EPU_OW, Hedge_bond_GPR_OW, Hedge_bond_CPU_OW, Hedge_bond_bs_OW],
    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)

wb0 = Workbook()
writer = pd.ExcelWriter(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_OW.xlsx', engine='openpyxl')
writer.book = wb0
OW_Gold.to_excel(writer, sheet_name='Gold')
OW_Oil.to_excel(writer, sheet_name='Oil')
OW_BTC.to_excel(writer, sheet_name='BTC')
OW_Bond.to_excel(writer, sheet_name='Bond')
wb0.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_OW.xlsx')

##**********************************************************************************************************************
# Hedge ratios 最优对冲比率
##**********************************************************************************************************************

## HR--CPU--------------------------------------------------------------------------------------------------------------
HR_gold_CPU = CoV_CPU_Gold / Var_CPU_Gold
HR_oil_CPU = CoV_CPU_Oil / Var_CPU_Oil
HR_btc_CPU = CoV_CPU_BTC / Var_CPU_BTC
HR_bond_CPU = CoV_CPU_Bond / Var_CPU_Bond

## HR--EMV--------------------------------------------------------------------------------------------------------------
HR_gold_EMV = CoV_EMV_Gold / Var_EMV_Gold
HR_oil_EMV = CoV_EMV_Oil / Var_EMV_Oil
HR_btc_EMV = CoV_EMV_BTC / Var_EMV_BTC
HR_bond_EMV = CoV_EMV_Bond / Var_EMV_Bond

## HR--EPU--------------------------------------------------------------------------------------------------------------
HR_gold_EPU = CoV_EPU_Gold / Var_EPU_Gold
HR_oil_EPU = CoV_EPU_Oil / Var_EPU_Oil
HR_btc_EPU = CoV_EPU_BTC / Var_EPU_BTC
HR_bond_EPU = CoV_EPU_Bond / Var_EPU_Bond

## HR--TPU--------------------------------------------------------------------------------------------------------------
HR_gold_TPU = CoV_TPU_Gold / Var_TPU_Gold
HR_oil_TPU = CoV_TPU_Oil / Var_TPU_Oil
HR_btc_TPU = CoV_TPU_BTC / Var_TPU_BTC
HR_bond_TPU = CoV_TPU_Bond / Var_TPU_Bond

## HR--GPR--------------------------------------------------------------------------------------------------------------
HR_gold_GPR = CoV_GPR_Gold / Var_GPR_Gold
HR_oil_GPR = CoV_GPR_Oil / Var_GPR_Oil
HR_btc_GPR = CoV_GPR_BTC / Var_GPR_BTC
HR_bond_GPR = CoV_GPR_Bond / Var_GPR_Bond

## HR--BS--------------------------------------------------------------------------------------------------------------
HR_gold_bs = CoV_bs_Gold / Var_bs_Gold
HR_oil_bs = CoV_bs_Oil / Var_bs_Oil
HR_btc_bs = CoV_bs_BTC / Var_bs_BTC
HR_bond_bs = CoV_bs_Bond / Var_bs_Bond

# -----------------------------------------------------------------------------------------------------------------------
HR_Gold = pd.concat([HR_gold_TPU, HR_gold_EMV, HR_gold_EPU, HR_gold_GPR, HR_gold_CPU, HR_gold_bs],
                    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
HR_Oil = pd.concat([HR_oil_TPU, HR_oil_EMV, HR_oil_EPU, HR_oil_GPR, HR_oil_CPU, HR_oil_bs],
                   keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
HR_BTC = pd.concat([HR_btc_TPU, HR_btc_EMV, HR_btc_EPU, HR_btc_GPR, HR_btc_CPU, HR_btc_bs],
                   keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
HR_Bond = pd.concat([HR_bond_TPU, HR_bond_EMV, HR_bond_EPU, HR_bond_GPR, HR_bond_CPU, HR_bond_bs],
                    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)

wb1 = Workbook()
writer = pd.ExcelWriter(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_OHR.xlsx', engine='openpyxl')
writer.book = wb1
HR_Gold.to_excel(writer, sheet_name='Gold')
HR_Oil.to_excel(writer, sheet_name='Oil')
HR_BTC.to_excel(writer, sheet_name='BTC')
HR_Bond.to_excel(writer, sheet_name='Bond')
wb1.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_OHR.xlsx')

##**********************************************************************************************************************
##**********************************************************************************************************************

# hedged Var--CPU
HVar_gold_CPU=np.square(HR_gold_CPU)*Var_CPU_Gold+Var_CPU_CSI300-2*HR_gold_CPU*CoV_CPU_Gold
HVar_oil_CPU=np.square(HR_oil_CPU)*Var_CPU_Oil+Var_CPU_CSI300-2*HR_oil_CPU*CoV_CPU_Oil
HVar_btc_CPU=np.square(HR_btc_CPU)*Var_CPU_BTC+Var_CPU_CSI300-2*HR_btc_CPU*CoV_CPU_BTC
HVar_bond_CPU=np.square(HR_bond_CPU)*Var_CPU_Bond+Var_CPU_CSI300-2*HR_bond_CPU*CoV_CPU_Bond

# ## hedged Var--EMV
HVar_gold_EMV=np.square(HR_gold_EMV)*Var_EMV_Gold+Var_EMV_CSI300-2*HR_gold_EMV*CoV_EMV_Gold
HVar_oil_EMV=np.square(HR_oil_EMV)*Var_EMV_Oil+Var_EMV_CSI300-2*HR_oil_EMV*CoV_EMV_Oil
HVar_btc_EMV=np.square(HR_btc_EMV)*Var_EMV_BTC+Var_EMV_CSI300-2*HR_btc_EMV*CoV_EMV_BTC
HVar_bond_EMV=np.square(HR_bond_EMV)*Var_EMV_Bond+Var_EMV_CSI300-2*HR_bond_EMV*CoV_EMV_Bond
#
# ## hedged Var--EPU
HVar_gold_EPU=np.square(HR_gold_EPU)*Var_EPU_Gold+Var_EPU_CSI300-2*HR_gold_EPU*CoV_EPU_Gold
HVar_oil_EPU=np.square(HR_oil_EPU)*Var_EPU_Oil+Var_EPU_CSI300-2*HR_oil_EPU*CoV_EPU_Oil
HVar_btc_EPU=np.square(HR_btc_EPU)*Var_EPU_BTC+Var_EPU_CSI300-2*HR_btc_EPU*CoV_EPU_BTC
HVar_bond_EPU=np.square(HR_bond_EPU)*Var_EPU_Bond+Var_EPU_CSI300-2*HR_bond_EPU*CoV_EPU_Bond
#
# ## hedged Var--TPU
HVar_gold_TPU=np.square(HR_gold_TPU)*Var_TPU_Gold+Var_TPU_CSI300-2*HR_gold_TPU*CoV_TPU_Gold
HVar_oil_TPU=np.square(HR_oil_TPU)*Var_TPU_Oil+Var_TPU_CSI300-2*HR_oil_TPU*CoV_TPU_Oil
HVar_btc_TPU=np.square(HR_btc_TPU)*Var_TPU_BTC+Var_TPU_CSI300-2*HR_btc_TPU*CoV_TPU_BTC
HVar_bond_TPU=np.square(HR_bond_TPU)*Var_TPU_Bond+Var_TPU_CSI300-2*HR_bond_TPU*CoV_TPU_Bond
#
# ## hedged Var--GPR
HVar_gold_GPR=np.square(HR_gold_GPR)*Var_GPR_Gold+Var_GPR_CSI300-2*HR_gold_GPR*CoV_GPR_Gold
HVar_oil_GPR=np.square(HR_oil_GPR)*Var_GPR_Oil+Var_GPR_CSI300-2*HR_oil_GPR*CoV_GPR_Oil
HVar_btc_GPR=np.square(HR_btc_GPR)*Var_GPR_BTC+Var_GPR_CSI300-2*HR_btc_GPR*CoV_GPR_BTC
HVar_bond_GPR=np.square(HR_bond_GPR)*Var_GPR_Bond+Var_GPR_CSI300-2*HR_bond_GPR*CoV_GPR_Bond

# ## hedged Var--BS
HVar_gold_bs=np.square(HR_gold_bs)*Var_bs_Gold+Var_bs_CSI300-2*HR_gold_bs*CoV_bs_Gold
HVar_oil_bs=np.square(HR_oil_bs)*Var_bs_Oil+Var_bs_CSI300-2*HR_oil_bs*CoV_bs_Oil
HVar_btc_bs=np.square(HR_btc_bs)*Var_bs_BTC+Var_bs_CSI300-2*HR_btc_bs*CoV_bs_BTC
HVar_bond_bs=np.square(HR_bond_bs)*Var_bs_Bond+Var_bs_CSI300-2*HR_bond_bs*CoV_bs_Bond

##**********************************************************************************************************************
# Hedge effectiveness 对冲效果
##**********************************************************************************************************************

##HE-CPU
HE_gold_CPU = 1 - (HVar_gold_CPU / Var_CPU_CSI300)
HE_oil_CPU = 1 - (HVar_oil_CPU / Var_CPU_CSI300)
HE_btc_CPU = 1 - (HVar_btc_CPU / Var_CPU_CSI300)
HE_bond_CPU = 1 - (HVar_bond_CPU / Var_CPU_CSI300)

##HE-EMV
HE_gold_EMV = 1 - (HVar_gold_EMV / Var_EMV_CSI300)
HE_oil_EMV = 1 - (HVar_oil_EMV / Var_EMV_CSI300)
HE_btc_EMV = 1 - (HVar_btc_EMV / Var_EMV_CSI300)
HE_bond_EMV = 1 - (HVar_bond_EMV / Var_EMV_CSI300)

##HE-EPU
HE_gold_EPU = 1 - (HVar_gold_EPU / Var_EPU_CSI300)
HE_oil_EPU = 1 - (HVar_oil_EPU / Var_EPU_CSI300)
HE_btc_EPU = 1 - (HVar_btc_EPU / Var_EPU_CSI300)
HE_bond_EPU = 1 - (HVar_bond_EPU / Var_EPU_CSI300)

##HE-TPU
HE_gold_TPU = 1 - (HVar_gold_TPU / Var_TPU_CSI300)
HE_oil_TPU = 1 - (HVar_oil_TPU / Var_TPU_CSI300)
HE_btc_TPU = 1 - (HVar_btc_TPU / Var_TPU_CSI300)
HE_bond_TPU = 1 - (HVar_bond_TPU / Var_TPU_CSI300)

##HE-GPR
HE_gold_GPR = 1 - (HVar_gold_GPR / Var_GPR_CSI300)
HE_oil_GPR = 1 - (HVar_oil_GPR / Var_GPR_CSI300)
HE_btc_GPR = 1 - (HVar_btc_GPR / Var_GPR_CSI300)
HE_bond_GPR = 1 - (HVar_bond_GPR / Var_GPR_CSI300)

##HE-BS
HE_gold_bs = 1 - (HVar_gold_bs / Var_bs_CSI300)
HE_oil_bs = 1 - (HVar_oil_bs / Var_bs_CSI300)
HE_btc_bs = 1 - (HVar_btc_bs / Var_bs_CSI300)
HE_bond_bs = 1 - (HVar_bond_bs / Var_bs_CSI300)

HE_Gold = pd.concat([HE_gold_TPU, HE_gold_EMV, HE_gold_EPU, HE_gold_GPR, HE_gold_CPU, HE_gold_bs],
                    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
HE_Oil = pd.concat([HE_oil_TPU, HE_oil_EMV, HE_oil_EPU, HE_oil_GPR, HE_oil_CPU, HE_oil_bs],
                   keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
HE_BTC = pd.concat([HE_btc_TPU, HE_btc_EMV, HE_btc_EPU, HE_btc_GPR, HE_btc_CPU, HE_btc_bs],
                   keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
HE_Bond = pd.concat([HE_bond_TPU, HE_bond_EMV, HE_bond_EPU, HE_bond_GPR, HE_bond_CPU, HE_bond_bs],
                    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)

wb2 = Workbook()
writer = pd.ExcelWriter(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_HE.xlsx', engine='openpyxl')
writer.book = wb2
HE_Gold.to_excel(writer, sheet_name='Gold')
HE_Oil.to_excel(writer, sheet_name='Oil')
HE_BTC.to_excel(writer, sheet_name='BTC')
HE_Bond.to_excel(writer, sheet_name='Bond')
wb2.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_HE.xlsx')
##**********************************************************************************************************************
##**********************************************************************************************************************


#
# return--CPU
rt_gold_CPU=StockCH-HR_gold_CPU*Gold
rt_oil_CPU=StockCH-HR_oil_CPU*Oil
rt_btc_CPU=StockCH-HR_btc_CPU*BTC
rt_bond_CPU=StockCH-HR_bond_CPU*Bond
# return--EMV
rt_gold_EMV=StockCH-HR_gold_EMV*Gold
rt_oil_EMV=StockCH-HR_oil_EMV*Oil
rt_btc_EMV=StockCH-HR_btc_EMV*BTC
rt_bond_EMV=StockCH-HR_bond_EMV*Bond
# return--EPU
rt_gold_EPU=StockCH-HR_gold_EPU*Gold
rt_oil_EPU=StockCH-HR_oil_EPU*Oil
rt_btc_EPU=StockCH-HR_btc_EPU*BTC
rt_bond_EPU=StockCH-HR_bond_EPU*Bond
# return--TPU
rt_gold_TPU=StockCH-HR_gold_TPU*Gold
rt_oil_TPU=StockCH-HR_oil_TPU*Oil
rt_btc_TPU=StockCH-HR_btc_TPU*BTC
rt_bond_TPU=StockCH-HR_bond_TPU*Bond
# return--GPR
rt_gold_GPR=StockCH-HR_gold_GPR*Gold
rt_oil_GPR=StockCH-HR_oil_GPR*Oil
rt_btc_GPR=StockCH-HR_btc_GPR*BTC
rt_bond_GPR=StockCH-HR_bond_GPR*Bond
#
# return--bs
rt_gold_bs=StockCH-HR_gold_bs*Gold
rt_oil_bs=StockCH-HR_oil_bs*Oil
rt_btc_bs=StockCH-HR_btc_bs*BTC
rt_bond_bs=StockCH-HR_bond_bs*Bond
##**********************************************************************************************************************
# Downside risk reduction
##**********************************************************************************************************************

z = norm.ppf(0.95)

#静态VaR & ES--CPU---------------------------------------------------------------------------------------------------------
SVaR_gold_CPU = qs.stats.value_at_risk(rt_gold_CPU)
SVaR_oil_CPU = qs.stats.value_at_risk(rt_oil_CPU)
SVaR_btc_CPU = qs.stats.value_at_risk(rt_btc_CPU)
SVaR_bond_CPU = qs.stats.value_at_risk(rt_bond_CPU)
SVaR_stock_CPU = qs.stats.value_at_risk(StockCH)

SES_gold_CPU=qs.stats.expected_shortfall(rt_gold_CPU)
SES_oil_CPU=qs.stats.expected_shortfall(rt_oil_CPU)
SES_btc_CPU=qs.stats.expected_shortfall(rt_btc_CPU)
SES_bond_CPU=qs.stats.expected_shortfall(rt_bond_CPU)
SES_stock_CPU = qs.stats.expected_shortfall(StockCH)

##LPM
LPM_gold_CPU=np.maximum(-rt_gold_CPU, 0)
LPM_oil_CPU=np.maximum(-rt_oil_CPU, 0)
LPM_btc_CPU=np.maximum(-rt_btc_CPU, 0)
LPM_bond_CPU=np.maximum(-rt_bond_CPU, 0)
LPM_stock_CPU=np.maximum(-StockCH, 0)

MeanLPM_gold_CPU=LPM_gold_CPU.mean()
MeanLPM_oil_CPU=LPM_oil_CPU.mean()
MeanLPM_btc_CPU=LPM_btc_CPU.mean()
MeanLPM_bond_CPU=LPM_bond_CPU.mean()
MeanLPM_stock_CPU=LPM_stock_CPU.mean()

#动态VaR & ES--CPU
DVaR_gold_CPU=rt_gold_CPU.mean()-z*np.sqrt(HVar_gold_CPU)
DVaR_oil_CPU=rt_oil_CPU.mean()-z*np.sqrt(HVar_oil_CPU)
DVaR_btc_CPU=rt_btc_CPU.mean()-z*np.sqrt(HVar_btc_CPU)
DVaR_bond_CPU=rt_bond_CPU.mean()-z*np.sqrt(HVar_bond_CPU)

VReduct_gold_CPU=1-(SVaR_gold_CPU/SVaR_stock_CPU)
VReduct_oil_CPU=1-(SVaR_oil_CPU/SVaR_stock_CPU)
VReduct_btc_CPU=1-(SVaR_btc_CPU/SVaR_stock_CPU)
VReduct_bond_CPU=1-(SVaR_bond_CPU/SVaR_stock_CPU)

ESReduct_gold_CPU=1-(SES_gold_CPU/SES_stock_CPU)
ESReduct_oil_CPU=1-(SES_oil_CPU/SES_stock_CPU)
ESReduct_btc_CPU=1-(SES_btc_CPU/SES_stock_CPU)
ESReduct_bond_CPU=1-(SES_bond_CPU/SES_stock_CPU)

LReduct_gold_CPU=1-(MeanLPM_gold_CPU/MeanLPM_stock_CPU)
LReduct_oil_CPU=1-(MeanLPM_oil_CPU/MeanLPM_stock_CPU)
LReduct_btc_CPU=1-(MeanLPM_btc_CPU/MeanLPM_stock_CPU)
LReduct_bond_CPU=1-(MeanLPM_bond_CPU/MeanLPM_stock_CPU)

#静态VaR & ES--EMV---------------------------------------------------------------------------------------------------------
SVaR_gold_EMV = qs.stats.value_at_risk(rt_gold_EMV)
SVaR_oil_EMV = qs.stats.value_at_risk(rt_oil_EMV)
SVaR_btc_EMV = qs.stats.value_at_risk(rt_btc_EMV)
SVaR_bond_EMV = qs.stats.value_at_risk(rt_bond_EMV)
SVaR_stock_EMV = qs.stats.value_at_risk(StockCH)

SES_gold_EMV=qs.stats.expected_shortfall(rt_gold_EMV)
SES_oil_EMV=qs.stats.expected_shortfall(rt_oil_EMV)
SES_btc_EMV=qs.stats.expected_shortfall(rt_btc_EMV)
SES_bond_EMV=qs.stats.expected_shortfall(rt_bond_EMV)
SES_stock_EMV = qs.stats.expected_shortfall(StockCH)

##LPM
LPM_gold_EMV=np.maximum(-rt_gold_EMV, 0)
LPM_oil_EMV=np.maximum(-rt_oil_EMV, 0)
LPM_btc_EMV=np.maximum(-rt_btc_EMV, 0)
LPM_bond_EMV=np.maximum(-rt_bond_EMV, 0)
LPM_stock_EMV=np.maximum(-StockCH, 0)

MeanLPM_gold_EMV=LPM_gold_EMV.mean()
MeanLPM_oil_EMV=LPM_oil_EMV.mean()
MeanLPM_btc_EMV=LPM_btc_EMV.mean()
MeanLPM_bond_EMV=LPM_bond_EMV.mean()
MeanLPM_stock_EMV=LPM_stock_EMV.mean()

#动态VaR & ES--EMV
DVaR_gold_EMV=rt_gold_EMV.mean()-z*np.sqrt(HVar_gold_EMV)
DVaR_oil_EMV=rt_oil_EMV.mean()-z*np.sqrt(HVar_oil_EMV)
DVaR_btc_EMV=rt_btc_EMV.mean()-z*np.sqrt(HVar_btc_EMV)
DVaR_bond_EMV=rt_bond_EMV.mean()-z*np.sqrt(HVar_bond_EMV)

VReduct_gold_EMV=1-(SVaR_gold_EMV/SVaR_stock_EMV)
VReduct_oil_EMV=1-(SVaR_oil_EMV/SVaR_stock_EMV)
VReduct_btc_EMV=1-(SVaR_btc_EMV/SVaR_stock_EMV)
VReduct_bond_EMV=1-(SVaR_bond_EMV/SVaR_stock_EMV)

ESReduct_gold_EMV=1-(SES_gold_EMV/SES_stock_EMV)
ESReduct_oil_EMV=1-(SES_oil_EMV/SES_stock_EMV)
ESReduct_btc_EMV=1-(SES_btc_EMV/SES_stock_EMV)
ESReduct_bond_EMV=1-(SES_bond_EMV/SES_stock_EMV)

LReduct_gold_EMV=1-(MeanLPM_gold_EMV/MeanLPM_stock_EMV)
LReduct_oil_EMV=1-(MeanLPM_oil_EMV/MeanLPM_stock_EMV)
LReduct_btc_EMV=1-(MeanLPM_btc_EMV/MeanLPM_stock_EMV)
LReduct_bond_EMV=1-(MeanLPM_bond_EMV/MeanLPM_stock_EMV)
#静态VaR & ES--EPU---------------------------------------------------------------------------------------------------------
SVaR_gold_EPU = qs.stats.value_at_risk(rt_gold_EPU)
SVaR_oil_EPU = qs.stats.value_at_risk(rt_oil_EPU)
SVaR_btc_EPU = qs.stats.value_at_risk(rt_btc_EPU)
SVaR_bond_EPU = qs.stats.value_at_risk(rt_bond_EPU)
SVaR_stock_EPU = qs.stats.value_at_risk(StockCH)

SES_gold_EPU=qs.stats.expected_shortfall(rt_gold_EPU)
SES_oil_EPU=qs.stats.expected_shortfall(rt_oil_EPU)
SES_btc_EPU=qs.stats.expected_shortfall(rt_btc_EPU)
SES_bond_EPU=qs.stats.expected_shortfall(rt_bond_EPU)
SES_stock_EPU = qs.stats.expected_shortfall(StockCH)

##LPM
LPM_gold_EPU=np.maximum(-rt_gold_EPU, 0)
LPM_oil_EPU=np.maximum(-rt_oil_EPU, 0)
LPM_btc_EPU=np.maximum(-rt_btc_EPU, 0)
LPM_bond_EPU=np.maximum(-rt_bond_EPU, 0)
LPM_stock_EPU=np.maximum(-StockCH, 0)

MeanLPM_gold_EPU=LPM_gold_EPU.mean()
MeanLPM_oil_EPU=LPM_oil_EPU.mean()
MeanLPM_btc_EPU=LPM_btc_EPU.mean()
MeanLPM_bond_EPU=LPM_bond_EPU.mean()
MeanLPM_stock_EPU=LPM_stock_EPU.mean()

#动态VaR & ES--EPU
DVaR_gold_EPU=rt_gold_EPU.mean()-z*np.sqrt(HVar_gold_EPU)
DVaR_oil_EPU=rt_oil_EPU.mean()-z*np.sqrt(HVar_oil_EPU)
DVaR_btc_EPU=rt_btc_EPU.mean()-z*np.sqrt(HVar_btc_EPU)
DVaR_bond_EPU=rt_bond_EPU.mean()-z*np.sqrt(HVar_bond_EPU)

VReduct_gold_EPU=1-(SVaR_gold_EPU/SVaR_stock_EPU)
VReduct_oil_EPU=1-(SVaR_oil_EPU/SVaR_stock_EPU)
VReduct_btc_EPU=1-(SVaR_btc_EPU/SVaR_stock_EPU)
VReduct_bond_EPU=1-(SVaR_bond_EPU/SVaR_stock_EPU)

ESReduct_gold_EPU=1-(SES_gold_EPU/SES_stock_EPU)
ESReduct_oil_EPU=1-(SES_oil_EPU/SES_stock_EPU)
ESReduct_btc_EPU=1-(SES_btc_EPU/SES_stock_EPU)
ESReduct_bond_EPU=1-(SES_bond_EPU/SES_stock_EPU)

LReduct_gold_EPU=1-(MeanLPM_gold_EPU/MeanLPM_stock_EPU)
LReduct_oil_EPU=1-(MeanLPM_oil_EPU/MeanLPM_stock_EPU)
LReduct_btc_EPU=1-(MeanLPM_btc_EPU/MeanLPM_stock_EPU)
LReduct_bond_EPU=1-(MeanLPM_bond_EPU/MeanLPM_stock_EPU)

#静态VaR & ES--TPU---------------------------------------------------------------------------------------------------------
SVaR_gold_TPU = qs.stats.value_at_risk(rt_gold_TPU)
SVaR_oil_TPU = qs.stats.value_at_risk(rt_oil_TPU)
SVaR_btc_TPU = qs.stats.value_at_risk(rt_btc_TPU)
SVaR_bond_TPU = qs.stats.value_at_risk(rt_bond_TPU)
SVaR_stock_TPU = qs.stats.value_at_risk(StockCH)

SES_gold_TPU=qs.stats.expected_shortfall(rt_gold_TPU)
SES_oil_TPU=qs.stats.expected_shortfall(rt_oil_TPU)
SES_btc_TPU=qs.stats.expected_shortfall(rt_btc_TPU)
SES_bond_TPU=qs.stats.expected_shortfall(rt_bond_TPU)
SES_stock_TPU = qs.stats.expected_shortfall(StockCH)

##LPM
LPM_gold_TPU=np.maximum(-rt_gold_TPU, 0)
LPM_oil_TPU=np.maximum(-rt_oil_TPU, 0)
LPM_btc_TPU=np.maximum(-rt_btc_TPU, 0)
LPM_bond_TPU=np.maximum(-rt_bond_TPU, 0)
LPM_stock_TPU=np.maximum(-StockCH, 0)

MeanLPM_gold_TPU=LPM_gold_TPU.mean()
MeanLPM_oil_TPU=LPM_oil_TPU.mean()
MeanLPM_btc_TPU=LPM_btc_TPU.mean()
MeanLPM_bond_TPU=LPM_bond_TPU.mean()
MeanLPM_stock_TPU=LPM_stock_TPU.mean()

#动态VaR & ES--TPU
DVaR_gold_TPU=rt_gold_TPU.mean()-z*np.sqrt(HVar_gold_TPU)
DVaR_oil_TPU=rt_oil_TPU.mean()-z*np.sqrt(HVar_oil_TPU)
DVaR_btc_TPU=rt_btc_TPU.mean()-z*np.sqrt(HVar_btc_TPU)
DVaR_bond_TPU=rt_bond_TPU.mean()-z*np.sqrt(HVar_bond_TPU)

VReduct_gold_TPU=1-(SVaR_gold_TPU/SVaR_stock_TPU)
VReduct_oil_TPU=1-(SVaR_oil_TPU/SVaR_stock_TPU)
VReduct_btc_TPU=1-(SVaR_btc_TPU/SVaR_stock_TPU)
VReduct_bond_TPU=1-(SVaR_bond_TPU/SVaR_stock_TPU)

ESReduct_gold_TPU=1-(SES_gold_TPU/SES_stock_TPU)
ESReduct_oil_TPU=1-(SES_oil_TPU/SES_stock_TPU)
ESReduct_btc_TPU=1-(SES_btc_TPU/SES_stock_TPU)
ESReduct_bond_TPU=1-(SES_bond_TPU/SES_stock_TPU)

LReduct_gold_TPU=1-(MeanLPM_gold_TPU/MeanLPM_stock_TPU)
LReduct_oil_TPU=1-(MeanLPM_oil_TPU/MeanLPM_stock_TPU)
LReduct_btc_TPU=1-(MeanLPM_btc_TPU/MeanLPM_stock_TPU)
LReduct_bond_TPU=1-(MeanLPM_bond_TPU/MeanLPM_stock_TPU)

#静态VaR & ES--GPR---------------------------------------------------------------------------------------------------------
SVaR_gold_GPR = qs.stats.value_at_risk(rt_gold_GPR)
SVaR_oil_GPR = qs.stats.value_at_risk(rt_oil_GPR)
SVaR_btc_GPR = qs.stats.value_at_risk(rt_btc_GPR)
SVaR_bond_GPR = qs.stats.value_at_risk(rt_bond_GPR)
SVaR_stock_GPR = qs.stats.value_at_risk(StockCH)

SES_gold_GPR=qs.stats.expected_shortfall(rt_gold_GPR)
SES_oil_GPR=qs.stats.expected_shortfall(rt_oil_GPR)
SES_btc_GPR=qs.stats.expected_shortfall(rt_btc_GPR)
SES_bond_GPR=qs.stats.expected_shortfall(rt_bond_GPR)
SES_stock_GPR = qs.stats.expected_shortfall(StockCH)

##LPM
LPM_gold_GPR=np.maximum(-rt_gold_GPR, 0)
LPM_oil_GPR=np.maximum(-rt_oil_GPR, 0)
LPM_btc_GPR=np.maximum(-rt_btc_GPR, 0)
LPM_bond_GPR=np.maximum(-rt_bond_GPR, 0)
LPM_stock_GPR=np.maximum(-StockCH, 0)

MeanLPM_gold_GPR=LPM_gold_GPR.mean()
MeanLPM_oil_GPR=LPM_oil_GPR.mean()
MeanLPM_btc_GPR=LPM_btc_GPR.mean()
MeanLPM_bond_GPR=LPM_bond_GPR.mean()
MeanLPM_stock_GPR=LPM_stock_GPR.mean()

#动态VaR & ES--GPR
DVaR_gold_GPR=rt_gold_GPR.mean()-z*np.sqrt(HVar_gold_GPR)
DVaR_oil_GPR=rt_oil_GPR.mean()-z*np.sqrt(HVar_oil_GPR)
DVaR_btc_GPR=rt_btc_GPR.mean()-z*np.sqrt(HVar_btc_GPR)
DVaR_bond_GPR=rt_bond_GPR.mean()-z*np.sqrt(HVar_bond_GPR)

VReduct_gold_GPR=1-(SVaR_gold_GPR/SVaR_stock_GPR)
VReduct_oil_GPR=1-(SVaR_oil_GPR/SVaR_stock_GPR)
VReduct_btc_GPR=1-(SVaR_btc_GPR/SVaR_stock_GPR)
VReduct_bond_GPR=1-(SVaR_bond_GPR/SVaR_stock_GPR)

ESReduct_gold_GPR=1-(SES_gold_GPR/SES_stock_GPR)
ESReduct_oil_GPR=1-(SES_oil_GPR/SES_stock_GPR)
ESReduct_btc_GPR=1-(SES_btc_GPR/SES_stock_GPR)
ESReduct_bond_GPR=1-(SES_bond_GPR/SES_stock_GPR)

LReduct_gold_GPR=1-(MeanLPM_gold_GPR/MeanLPM_stock_GPR)
LReduct_oil_GPR=1-(MeanLPM_oil_GPR/MeanLPM_stock_GPR)
LReduct_btc_GPR=1-(MeanLPM_btc_GPR/MeanLPM_stock_GPR)
LReduct_bond_GPR=1-(MeanLPM_bond_GPR/MeanLPM_stock_GPR)


#静态VaR & ES--bs---------------------------------------------------------------------------------------------------------
SVaR_gold_bs = qs.stats.value_at_risk(rt_gold_bs)
SVaR_oil_bs = qs.stats.value_at_risk(rt_oil_bs)
SVaR_btc_bs = qs.stats.value_at_risk(rt_btc_bs)
SVaR_bond_bs = qs.stats.value_at_risk(rt_bond_bs)
SVaR_stock_bs = qs.stats.value_at_risk(StockCH)

SES_gold_bs=qs.stats.expected_shortfall(rt_gold_bs)
SES_oil_bs=qs.stats.expected_shortfall(rt_oil_bs)
SES_btc_bs=qs.stats.expected_shortfall(rt_btc_bs)
SES_bond_bs=qs.stats.expected_shortfall(rt_bond_bs)
SES_stock_bs = qs.stats.expected_shortfall(StockCH)

##LPM
LPM_gold_bs=np.maximum(-rt_gold_bs, 0)
LPM_oil_bs=np.maximum(-rt_oil_bs, 0)
LPM_btc_bs=np.maximum(-rt_btc_bs, 0)
LPM_bond_bs=np.maximum(-rt_bond_bs, 0)
LPM_stock_bs=np.maximum(-StockCH, 0)

MeanLPM_gold_bs=LPM_gold_bs.mean()
MeanLPM_oil_bs=LPM_oil_bs.mean()
MeanLPM_btc_bs=LPM_btc_bs.mean()
MeanLPM_bond_bs=LPM_bond_bs.mean()
MeanLPM_stock_bs=LPM_stock_bs.mean()

#动态VaR & ES--bs
DVaR_gold_bs=rt_gold_bs.mean()-z*np.sqrt(HVar_gold_bs)
DVaR_oil_bs=rt_oil_bs.mean()-z*np.sqrt(HVar_oil_bs)
DVaR_btc_bs=rt_btc_bs.mean()-z*np.sqrt(HVar_btc_bs)
DVaR_bond_bs=rt_bond_bs.mean()-z*np.sqrt(HVar_bond_bs)

VReduct_gold_bs=1-(SVaR_gold_bs/SVaR_stock_bs)
VReduct_oil_bs=1-(SVaR_oil_bs/SVaR_stock_bs)
VReduct_btc_bs=1-(SVaR_btc_bs/SVaR_stock_bs)
VReduct_bond_bs=1-(SVaR_bond_bs/SVaR_stock_bs)

ESReduct_gold_bs=1-(SES_gold_bs/SES_stock_bs)
ESReduct_oil_bs=1-(SES_oil_bs/SES_stock_bs)
ESReduct_btc_bs=1-(SES_btc_bs/SES_stock_bs)
ESReduct_bond_bs=1-(SES_bond_bs/SES_stock_bs)

LReduct_gold_bs=1-(MeanLPM_gold_bs/MeanLPM_stock_bs)
LReduct_oil_bs=1-(MeanLPM_oil_bs/MeanLPM_stock_bs)
LReduct_btc_bs=1-(MeanLPM_btc_bs/MeanLPM_stock_bs)
LReduct_bond_bs=1-(MeanLPM_bond_bs/MeanLPM_stock_bs)
##**********************************************************************************************************************
##**********************************************************************************************************************


DVaR_Gold = pd.concat([DVaR_gold_TPU, DVaR_gold_EMV, DVaR_gold_EPU, DVaR_gold_GPR, DVaR_gold_CPU, DVaR_gold_bs],
                      keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
DVaR_Oil = pd.concat([DVaR_oil_TPU, DVaR_oil_EMV, DVaR_oil_EPU, DVaR_oil_GPR, DVaR_oil_CPU, DVaR_oil_bs],
                     keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
DVaR_BTC = pd.concat([DVaR_btc_TPU, DVaR_btc_EMV, DVaR_btc_EPU, DVaR_btc_GPR, DVaR_btc_CPU, DVaR_btc_bs],
                     keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
DVaR_Bond = pd.concat([DVaR_bond_TPU, DVaR_bond_EMV, DVaR_bond_EPU, DVaR_bond_GPR, DVaR_bond_CPU, DVaR_bond_bs],
                      keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)

wb3 = Workbook()
writer = pd.ExcelWriter(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_VaR.xlsx', engine='openpyxl')
writer.book = wb3
DVaR_Gold.to_excel(writer, sheet_name='Gold')
DVaR_Oil.to_excel(writer, sheet_name='Oil')
DVaR_BTC.to_excel(writer, sheet_name='BTC')
DVaR_Bond.to_excel(writer, sheet_name='Bond')
wb3.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_VaR.xlsx')

LPM_Gold = pd.concat([LPM_gold_TPU, LPM_gold_EMV, LPM_gold_EPU, LPM_gold_GPR, LPM_gold_CPU, LPM_gold_bs],
                     keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
LPM_Oil = pd.concat([LPM_oil_TPU, LPM_oil_EMV, LPM_oil_EPU, LPM_oil_GPR, LPM_oil_CPU, LPM_oil_bs],
                    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
LPM_BTC = pd.concat([LPM_btc_TPU, LPM_btc_EMV, LPM_btc_EPU, LPM_btc_GPR, LPM_btc_CPU, LPM_btc_bs],
                    keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)
LPM_Bond = pd.concat([LPM_bond_TPU, LPM_bond_EMV, LPM_bond_EPU, LPM_bond_GPR, LPM_bond_CPU, LPM_bond_bs],
                     keys=['TPU', 'EMV', 'EPU', 'GPR', 'CPU', 'Baseline'], axis=1)

wb4 = Workbook()
writer = pd.ExcelWriter(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_LPM.xlsx', engine='openpyxl')
writer.book = wb4
LPM_Gold.to_excel(writer, sheet_name='Gold')
LPM_Oil.to_excel(writer, sheet_name='Oil')
LPM_BTC.to_excel(writer, sheet_name='BTC')
LPM_Bond.to_excel(writer, sheet_name='Bond')
wb4.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_LPM.xlsx')

##**********************************************************************************************************************
##**********************************************************************************************************************


##
## Diversification--CPU----------------------------------------------------------------------------------------------------
## Sharpe
Sharpe_gold_CPU = sharpe_ratio(rt_gold_CPU)
Sharpe_oil_CPU = sharpe_ratio(rt_oil_CPU)
Sharpe_btc_CPU = sharpe_ratio(rt_btc_CPU)
Sharpe_bond_CPU = sharpe_ratio(rt_bond_CPU)

# Sortino
Sortino_gold_CPU = sortino_ratio(rt_gold_CPU)
Sortino_oil_CPU = sortino_ratio(rt_oil_CPU)
Sortino_btc_CPU = sortino_ratio(rt_btc_CPU)
Sortino_bond_CPU = sortino_ratio(rt_bond_CPU)

# Omega
Omega_gold_CPU = omega_ratio(rt_gold_CPU)
Omega_oil_CPU = omega_ratio(rt_oil_CPU)
Omega_btc_CPU = omega_ratio(rt_btc_CPU)
Omega_bond_CPU = omega_ratio(rt_bond_CPU)

# Calmar
Calmar_gold_CPU = calmar_ratio(rt_gold_CPU)
Calmar_oil_CPU = calmar_ratio(rt_oil_CPU)
Calmar_btc_CPU = calmar_ratio(rt_btc_CPU)
Calmar_bond_CPU = calmar_ratio(rt_bond_CPU)

## Diversification--EMV----------------------------------------------------------------------------------------------------
## Sharpe
Sharpe_gold_EMV = sharpe_ratio(rt_gold_EMV)
Sharpe_oil_EMV = sharpe_ratio(rt_oil_EMV)
Sharpe_btc_EMV = sharpe_ratio(rt_btc_EMV)
Sharpe_bond_EMV = sharpe_ratio(rt_bond_EMV)

# Sortino
Sortino_gold_EMV = sortino_ratio(rt_gold_EMV)
Sortino_oil_EMV = sortino_ratio(rt_oil_EMV)
Sortino_btc_EMV = sortino_ratio(rt_btc_EMV)
Sortino_bond_EMV = sortino_ratio(rt_bond_EMV)

# Omega
Omega_gold_EMV = omega_ratio(rt_gold_EMV)
Omega_oil_EMV = omega_ratio(rt_oil_EMV)
Omega_btc_EMV = omega_ratio(rt_btc_EMV)
Omega_bond_EMV = omega_ratio(rt_bond_EMV)

# Calmar
Calmar_gold_EMV = calmar_ratio(rt_gold_EMV)
Calmar_oil_EMV = calmar_ratio(rt_oil_EMV)
Calmar_btc_EMV = calmar_ratio(rt_btc_EMV)
Calmar_bond_EMV = calmar_ratio(rt_bond_EMV)

## Diversification--EPU----------------------------------------------------------------------------------------------------
## Sharpe
Sharpe_gold_EPU = sharpe_ratio(rt_gold_EPU)
Sharpe_oil_EPU = sharpe_ratio(rt_oil_EPU)
Sharpe_btc_EPU = sharpe_ratio(rt_btc_EPU)
Sharpe_bond_EPU = sharpe_ratio(rt_bond_EPU)

# Sortino
Sortino_gold_EPU = sortino_ratio(rt_gold_EPU)
Sortino_oil_EPU = sortino_ratio(rt_oil_EPU)
Sortino_btc_EPU = sortino_ratio(rt_btc_EPU)
Sortino_bond_EPU = sortino_ratio(rt_bond_EPU)

# Omega
Omega_gold_EPU = omega_ratio(rt_gold_EPU)
Omega_oil_EPU = omega_ratio(rt_oil_EPU)
Omega_btc_EPU = omega_ratio(rt_btc_EPU)
Omega_bond_EPU = omega_ratio(rt_bond_EPU)

# Calmar
Calmar_gold_EPU = calmar_ratio(rt_gold_EPU)
Calmar_oil_EPU = calmar_ratio(rt_oil_EPU)
Calmar_btc_EPU = calmar_ratio(rt_btc_EPU)
Calmar_bond_EPU = calmar_ratio(rt_bond_EPU)

## Diversification--TPU----------------------------------------------------------------------------------------------------
## Sharpe
Sharpe_gold_TPU = sharpe_ratio(rt_gold_TPU)
Sharpe_oil_TPU = sharpe_ratio(rt_oil_TPU)
Sharpe_btc_TPU = sharpe_ratio(rt_btc_TPU)
Sharpe_bond_TPU = sharpe_ratio(rt_bond_TPU)

# Sortino
Sortino_gold_TPU = sortino_ratio(rt_gold_TPU)
Sortino_oil_TPU = sortino_ratio(rt_oil_TPU)
Sortino_btc_TPU = sortino_ratio(rt_btc_TPU)
Sortino_bond_TPU = sortino_ratio(rt_bond_TPU)

# Omega
Omega_gold_TPU = omega_ratio(rt_gold_TPU)
Omega_oil_TPU = omega_ratio(rt_oil_TPU)
Omega_btc_TPU = omega_ratio(rt_btc_TPU)
Omega_bond_TPU = omega_ratio(rt_bond_TPU)

# Calmar
Calmar_gold_TPU = calmar_ratio(rt_gold_TPU)
Calmar_oil_TPU = calmar_ratio(rt_oil_TPU)
Calmar_btc_TPU = calmar_ratio(rt_btc_TPU)
Calmar_bond_TPU = calmar_ratio(rt_bond_TPU)

## Diversification--GPR----------------------------------------------------------------------------------------------------
## Sharpe
Sharpe_gold_GPR = sharpe_ratio(rt_gold_GPR)
Sharpe_oil_GPR = sharpe_ratio(rt_oil_GPR)
Sharpe_btc_GPR = sharpe_ratio(rt_btc_GPR)
Sharpe_bond_GPR = sharpe_ratio(rt_bond_GPR)

# Sortino
Sortino_gold_GPR = sortino_ratio(rt_gold_GPR)
Sortino_oil_GPR = sortino_ratio(rt_oil_GPR)
Sortino_btc_GPR = sortino_ratio(rt_btc_GPR)
Sortino_bond_GPR = sortino_ratio(rt_bond_GPR)

# Omega
Omega_gold_GPR = omega_ratio(rt_gold_GPR)
Omega_oil_GPR = omega_ratio(rt_oil_GPR)
Omega_btc_GPR = omega_ratio(rt_btc_GPR)
Omega_bond_GPR = omega_ratio(rt_bond_GPR)

# Calmar
Calmar_gold_GPR = calmar_ratio(rt_gold_GPR)
Calmar_oil_GPR = calmar_ratio(rt_oil_GPR)
Calmar_btc_GPR = calmar_ratio(rt_btc_GPR)
Calmar_bond_GPR = calmar_ratio(rt_bond_GPR)

## Diversification--bs----------------------------------------------------------------------------------------------------
## Sharpe
Sharpe_gold_bs = sharpe_ratio(rt_gold_bs)
Sharpe_oil_bs = sharpe_ratio(rt_oil_bs)
Sharpe_btc_bs = sharpe_ratio(rt_btc_bs)
Sharpe_bond_bs = sharpe_ratio(rt_bond_bs)

# Sortino
Sortino_gold_bs = sortino_ratio(rt_gold_bs)
Sortino_oil_bs = sortino_ratio(rt_oil_bs)
Sortino_btc_bs = sortino_ratio(rt_btc_bs)
Sortino_bond_bs = sortino_ratio(rt_bond_bs)

# Omega
Omega_gold_bs = omega_ratio(rt_gold_bs)
Omega_oil_bs = omega_ratio(rt_oil_bs)
Omega_btc_bs = omega_ratio(rt_btc_bs)
Omega_bond_bs = omega_ratio(rt_bond_bs)

# Calmar
Calmar_gold_bs = calmar_ratio(rt_gold_bs)
Calmar_oil_bs = calmar_ratio(rt_oil_bs)
Calmar_btc_bs = calmar_ratio(rt_btc_bs)
Calmar_bond_bs = calmar_ratio(rt_bond_bs)

##***********************************************************************************************************************
##***********************************************************************************************************************

wb5 = Workbook()
ws5 = wb5.create_sheet("Diversification")
for index, value in enumerate(['EPU', 'CPU', 'TPU', 'EMV', 'GPR', 'Baseline'], start=2):
    ws5.cell(row=2, column=index, value=value)
ws5['A3'] = 'VaR reduction'
ws5['A8'] = 'ES reduction'
ws5['A13'] = 'LPM reduction'
ws5['A18'] = 'Sharpe ratio'
ws5['A23'] = 'Sortino ratio'
ws5['A28'] = 'Omega ratio'
ws5['A33'] = 'Calmar ratio'
for index, value in enumerate(
        ['Bond', VReduct_bond_EPU, VReduct_bond_CPU, VReduct_bond_TPU, VReduct_bond_EMV, VReduct_bond_GPR,
         VReduct_bond_bs], start=1):
    ws5.cell(row=4, column=index, value=value)
for index, value in enumerate(
        ['Gold', VReduct_gold_EPU, VReduct_gold_CPU, VReduct_gold_TPU, VReduct_gold_EMV, VReduct_gold_GPR,
         VReduct_gold_bs], start=1):
    ws5.cell(row=5, column=index, value=value)
for index, value in enumerate(
        ['Oil', VReduct_oil_EPU, VReduct_oil_CPU, VReduct_oil_TPU, VReduct_oil_EMV, VReduct_oil_GPR, VReduct_oil_bs],
        start=1):
    ws5.cell(row=6, column=index, value=value)
for index, value in enumerate(
        ['Bitcoin', VReduct_btc_EPU, VReduct_btc_CPU, VReduct_btc_TPU, VReduct_btc_EMV, VReduct_btc_GPR,
         VReduct_btc_bs], start=1):
    ws5.cell(row=7, column=index, value=value)

for index, value in enumerate(
        ['Bond', ESReduct_bond_EPU, ESReduct_bond_CPU, ESReduct_bond_TPU, ESReduct_bond_EMV, ESReduct_bond_GPR,
         ESReduct_bond_bs], start=1):
    ws5.cell(row=9, column=index, value=value)
for index, value in enumerate(
        ['Gold', ESReduct_gold_EPU, ESReduct_gold_CPU, ESReduct_gold_TPU, ESReduct_gold_EMV, ESReduct_gold_GPR,
         ESReduct_gold_bs], start=1):
    ws5.cell(row=10, column=index, value=value)
for index, value in enumerate(
        ['Oil', ESReduct_oil_EPU, ESReduct_oil_CPU, ESReduct_oil_TPU, ESReduct_oil_EMV, ESReduct_oil_GPR,
         ESReduct_oil_bs], start=1):
    ws5.cell(row=11, column=index, value=value)
for index, value in enumerate(
        ['Bitcoin', ESReduct_btc_EPU, ESReduct_btc_CPU, ESReduct_btc_TPU, ESReduct_btc_EMV, ESReduct_btc_GPR,
         ESReduct_btc_bs], start=1):
    ws5.cell(row=12, column=index, value=value)

for index, value in enumerate(
        ['Bond', LReduct_bond_EPU, LReduct_bond_CPU, LReduct_bond_TPU, LReduct_bond_EMV, LReduct_bond_GPR,
         LReduct_bond_bs], start=1):
    ws5.cell(row=14, column=index, value=value)
for index, value in enumerate(
        ['Gold', LReduct_gold_EPU, LReduct_gold_CPU, LReduct_gold_TPU, LReduct_gold_EMV, LReduct_gold_GPR,
         LReduct_gold_bs], start=1):
    ws5.cell(row=15, column=index, value=value)
for index, value in enumerate(
        ['Oil', LReduct_oil_EPU, LReduct_oil_CPU, LReduct_oil_TPU, LReduct_oil_EMV, LReduct_oil_GPR, LReduct_oil_bs],
        start=1):
    ws5.cell(row=16, column=index, value=value)
for index, value in enumerate(
        ['Bitcoin', LReduct_btc_EPU, LReduct_btc_CPU, LReduct_btc_TPU, LReduct_btc_EMV, LReduct_btc_GPR,
         LReduct_btc_bs], start=1):
    ws5.cell(row=17, column=index, value=value)

##**********************************************************************************************************************
##**********************************************************************************************************************


for index, value in enumerate(
        ['Bond', Sharpe_bond_EPU, Sharpe_bond_CPU, Sharpe_bond_TPU, Sharpe_bond_EMV, Sharpe_bond_GPR, Sharpe_bond_bs],
        start=1):
    ws5.cell(row=19, column=index, value=value)
for index, value in enumerate(
        ['Gold', Sharpe_gold_EPU, Sharpe_gold_CPU, Sharpe_gold_TPU, Sharpe_gold_EMV, Sharpe_gold_GPR, Sharpe_gold_bs],
        start=1):
    ws5.cell(row=20, column=index, value=value)
for index, value in enumerate(
        ['Oil', Sharpe_oil_EPU, Sharpe_oil_CPU, Sharpe_oil_TPU, Sharpe_oil_EMV, Sharpe_oil_GPR, Sharpe_oil_bs],
        start=1):
    ws5.cell(row=21, column=index, value=value)
for index, value in enumerate(
        ['Bitcoin', Sharpe_btc_EPU, Sharpe_btc_CPU, Sharpe_btc_TPU, Sharpe_btc_EMV, Sharpe_btc_GPR, Sharpe_btc_bs],
        start=1):
    ws5.cell(row=22, column=index, value=value)

for index, value in enumerate(
        ['Bond', Sortino_bond_EPU, Sortino_bond_CPU, Sortino_bond_TPU, Sortino_bond_EMV, Sortino_bond_GPR,
         Sortino_bond_bs], start=1):
    ws5.cell(row=24, column=index, value=value)
for index, value in enumerate(
        ['Gold', Sortino_gold_EPU, Sortino_gold_CPU, Sortino_gold_TPU, Sortino_gold_EMV, Sortino_gold_GPR,
         Sortino_gold_bs], start=1):
    ws5.cell(row=25, column=index, value=value)
for index, value in enumerate(
        ['Oil', Sortino_oil_EPU, Sortino_oil_CPU, Sortino_oil_TPU, Sortino_oil_EMV, Sortino_oil_GPR, Sortino_oil_bs],
        start=1):
    ws5.cell(row=26, column=index, value=value)
for index, value in enumerate(
        ['Bitcoin', Sortino_btc_EPU, Sortino_btc_CPU, Sortino_btc_TPU, Sortino_btc_EMV, Sortino_btc_GPR,
         Sortino_btc_bs], start=1):
    ws5.cell(row=27, column=index, value=value)

for index, value in enumerate(
        ['Bond', Omega_bond_EPU, Omega_bond_CPU, Omega_bond_TPU, Omega_bond_EMV, Omega_bond_GPR, Omega_bond_bs],
        start=1):
    ws5.cell(row=29, column=index, value=value)
for index, value in enumerate(
        ['Gold', Omega_gold_EPU, Omega_gold_CPU, Omega_gold_TPU, Omega_gold_EMV, Omega_gold_GPR, Omega_gold_bs],
        start=1):
    ws5.cell(row=30, column=index, value=value)
for index, value in enumerate(
        ['Oil', Omega_oil_EPU, Omega_oil_CPU, Omega_oil_TPU, Omega_oil_EMV, Omega_oil_GPR, Omega_oil_bs], start=1):
    ws5.cell(row=31, column=index, value=value)
for index, value in enumerate(
        ['Bitcoin', Omega_btc_EPU, Omega_btc_CPU, Omega_btc_TPU, Omega_btc_EMV, Omega_btc_GPR, Omega_btc_bs], start=1):
    ws5.cell(row=32, column=index, value=value)

for index, value in enumerate(
        ['Bond', Calmar_bond_EPU, Calmar_bond_CPU, Calmar_bond_TPU, Calmar_bond_EMV, Calmar_bond_GPR, Calmar_bond_bs],
        start=1):
    ws5.cell(row=34, column=index, value=value)
for index, value in enumerate(
        ['Gold', Calmar_gold_EPU, Calmar_gold_CPU, Calmar_gold_TPU, Calmar_gold_EMV, Calmar_gold_GPR, Calmar_gold_bs],
        start=1):
    ws5.cell(row=35, column=index, value=value)
for index, value in enumerate(
        ['Oil', Calmar_oil_EPU, Calmar_oil_CPU, Calmar_oil_TPU, Calmar_oil_EMV, Calmar_oil_GPR, Calmar_oil_bs],
        start=1):
    ws5.cell(row=36, column=index, value=value)
for index, value in enumerate(
        ['Bitcoin', Calmar_btc_EPU, Calmar_btc_CPU, Calmar_btc_TPU, Calmar_btc_EMV, Calmar_btc_GPR, Calmar_btc_bs],
        start=1):
    ws5.cell(row=37, column=index, value=value)

wb5.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_Diversification.xlsx')

##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************


wb6 = Workbook()
ws6 = wb6.create_sheet("Statistics")

for index, value in enumerate(['Mean', 'Max.', 'Min.', 'Std. Dev.'] * 3, start=2):
    ws6.cell(row=2, column=index, value=value)
ws6['B1'] = 'Optimal weights'
ws6['F1'] = 'Hedge ratios'
ws6['J1'] = 'Hedging effectiveness'

ws6['A3'] = 'Bond'
ws6['A10'] = 'Gold'
ws6['A17'] = 'Oil'
ws6['A24'] = 'Bitcoin'

for index, value in enumerate(
        ['EPU', Hedge_bond_EPU_OW.mean(), Hedge_bond_EPU_OW.max(), Hedge_bond_EPU_OW.min(), Hedge_bond_EPU_OW.std(),
         HR_bond_EPU.mean(), HR_bond_EPU.max(), HR_bond_EPU.min(), HR_bond_EPU.std(),
         HE_bond_EPU.mean(), HE_bond_EPU.max(), HE_bond_EPU.min(), HE_bond_EPU.std()], start=1):
    ws6.cell(row=4, column=index, value=value)
for index, value in enumerate(
        ['CPU', Hedge_bond_CPU_OW.mean(), Hedge_bond_CPU_OW.max(), Hedge_bond_CPU_OW.min(), Hedge_bond_CPU_OW.std(),
         HR_bond_CPU.mean(), HR_bond_CPU.max(), HR_bond_CPU.min(), HR_bond_CPU.std(),
         HE_bond_CPU.mean(), HE_bond_CPU.max(), HE_bond_CPU.min(), HE_bond_CPU.std()], start=1):
    ws6.cell(row=5, column=index, value=value)
for index, value in enumerate(
        ['TPU', Hedge_bond_TPU_OW.mean(), Hedge_bond_TPU_OW.max(), Hedge_bond_TPU_OW.min(), Hedge_bond_TPU_OW.std(),
         HR_bond_TPU.mean(), HR_bond_TPU.max(), HR_bond_TPU.min(), HR_bond_TPU.std(),
         HE_bond_TPU.mean(), HE_bond_TPU.max(), HE_bond_TPU.min(), HE_bond_TPU.std()], start=1):
    ws6.cell(row=6, column=index, value=value)
for index, value in enumerate(
        ['EMV', Hedge_bond_EMV_OW.mean(), Hedge_bond_EMV_OW.max(), Hedge_bond_EMV_OW.min(), Hedge_bond_EMV_OW.std(),
         HR_bond_EMV.mean(), HR_bond_EMV.max(), HR_bond_EMV.min(), HR_bond_EMV.std(),
         HE_bond_EMV.mean(), HE_bond_EMV.max(), HE_bond_EMV.min(), HE_bond_EMV.std()], start=1):
    ws6.cell(row=7, column=index, value=value)
for index, value in enumerate(
        ['GPR', Hedge_bond_GPR_OW.mean(), Hedge_bond_GPR_OW.max(), Hedge_bond_GPR_OW.min(), Hedge_bond_GPR_OW.std(),
         HR_bond_GPR.mean(), HR_bond_GPR.max(), HR_bond_GPR.min(), HR_bond_GPR.std(),
         HE_bond_GPR.mean(), HE_bond_GPR.max(), HE_bond_GPR.min(), HE_bond_GPR.std()], start=1):
    ws6.cell(row=8, column=index, value=value)
for index, value in enumerate(
        ['Baseline', Hedge_bond_bs_OW.mean(), Hedge_bond_bs_OW.max(), Hedge_bond_bs_OW.min(), Hedge_bond_bs_OW.std(),
         HR_bond_bs.mean(), HR_bond_bs.max(), HR_bond_bs.min(), HR_bond_bs.std(),
         HE_bond_bs.mean(), HE_bond_bs.max(), HE_bond_bs.min(), HE_bond_bs.std()], start=1):
    ws6.cell(row=9, column=index, value=value)

for index, value in enumerate(
        ['EPU', Hedge_gold_EPU_OW.mean(), Hedge_gold_EPU_OW.max(), Hedge_gold_EPU_OW.min(), Hedge_gold_EPU_OW.std(),
         HR_gold_EPU.mean(), HR_gold_EPU.max(), HR_gold_EPU.min(), HR_gold_EPU.std(),
         HE_gold_EPU.mean(), HE_gold_EPU.max(), HE_gold_EPU.min(), HE_gold_EPU.std()], start=1):
    ws6.cell(row=11, column=index, value=value)
for index, value in enumerate(
        ['CPU', Hedge_gold_CPU_OW.mean(), Hedge_gold_CPU_OW.max(), Hedge_gold_CPU_OW.min(), Hedge_gold_CPU_OW.std(),
         HR_gold_CPU.mean(), HR_gold_CPU.max(), HR_gold_CPU.min(), HR_gold_CPU.std(),
         HE_gold_CPU.mean(), HE_gold_CPU.max(), HE_gold_CPU.min(), HE_gold_CPU.std()], start=1):
    ws6.cell(row=12, column=index, value=value)
for index, value in enumerate(
        ['TPU', Hedge_gold_TPU_OW.mean(), Hedge_gold_TPU_OW.max(), Hedge_gold_TPU_OW.min(), Hedge_gold_TPU_OW.std(),
         HR_gold_TPU.mean(), HR_gold_TPU.max(), HR_gold_TPU.min(), HR_gold_TPU.std(),
         HE_gold_TPU.mean(), HE_gold_TPU.max(), HE_gold_TPU.min(), HE_gold_TPU.std()], start=1):
    ws6.cell(row=13, column=index, value=value)
for index, value in enumerate(
        ['EMV', Hedge_gold_EMV_OW.mean(), Hedge_gold_EMV_OW.max(), Hedge_gold_EMV_OW.min(), Hedge_gold_EMV_OW.std(),
         HR_gold_EMV.mean(), HR_gold_EMV.max(), HR_gold_EMV.min(), HR_gold_EMV.std(),
         HE_gold_EMV.mean(), HE_gold_EMV.max(), HE_gold_EMV.min(), HE_gold_EMV.std()], start=1):
    ws6.cell(row=14, column=index, value=value)
for index, value in enumerate(
        ['GPR', Hedge_gold_GPR_OW.mean(), Hedge_gold_GPR_OW.max(), Hedge_gold_GPR_OW.min(), Hedge_gold_GPR_OW.std(),
         HR_gold_GPR.mean(), HR_gold_GPR.max(), HR_gold_GPR.min(), HR_gold_GPR.std(),
         HE_gold_GPR.mean(), HE_gold_GPR.max(), HE_gold_GPR.min(), HE_gold_GPR.std()], start=1):
    ws6.cell(row=15, column=index, value=value)
for index, value in enumerate(
        ['Baseline', Hedge_gold_bs_OW.mean(), Hedge_gold_bs_OW.max(), Hedge_gold_bs_OW.min(), Hedge_gold_bs_OW.std(),
         HR_gold_bs.mean(), HR_gold_bs.max(), HR_gold_bs.min(), HR_gold_bs.std(),
         HE_gold_bs.mean(), HE_gold_bs.max(), HE_gold_bs.min(), HE_gold_bs.std()], start=1):
    ws6.cell(row=16, column=index, value=value)

for index, value in enumerate(
        ['EPU', Hedge_oil_EPU_OW.mean(), Hedge_oil_EPU_OW.max(), Hedge_oil_EPU_OW.min(), Hedge_oil_EPU_OW.std(),
         HR_oil_EPU.mean(), HR_oil_EPU.max(), HR_oil_EPU.min(), HR_oil_EPU.std(),
         HE_oil_EPU.mean(), HE_oil_EPU.max(), HE_oil_EPU.min(), HE_oil_EPU.std()], start=1):
    ws6.cell(row=18, column=index, value=value)
for index, value in enumerate(
        ['CPU', Hedge_oil_CPU_OW.mean(), Hedge_oil_CPU_OW.max(), Hedge_oil_CPU_OW.min(), Hedge_oil_CPU_OW.std(),
         HR_oil_CPU.mean(), HR_oil_CPU.max(), HR_oil_CPU.min(), HR_oil_CPU.std(),
         HE_oil_CPU.mean(), HE_oil_CPU.max(), HE_oil_CPU.min(), HE_oil_CPU.std()], start=1):
    ws6.cell(row=19, column=index, value=value)
for index, value in enumerate(
        ['TPU', Hedge_oil_TPU_OW.mean(), Hedge_oil_TPU_OW.max(), Hedge_oil_TPU_OW.min(), Hedge_oil_TPU_OW.std(),
         HR_oil_TPU.mean(), HR_oil_TPU.max(), HR_oil_TPU.min(), HR_oil_TPU.std(),
         HE_oil_TPU.mean(), HE_oil_TPU.max(), HE_oil_TPU.min(), HE_oil_TPU.std()], start=1):
    ws6.cell(row=20, column=index, value=value)
for index, value in enumerate(
        ['EMV', Hedge_oil_EMV_OW.mean(), Hedge_oil_EMV_OW.max(), Hedge_oil_EMV_OW.min(), Hedge_oil_EMV_OW.std(),
         HR_oil_EMV.mean(), HR_oil_EMV.max(), HR_oil_EMV.min(), HR_oil_EMV.std(),
         HE_oil_EMV.mean(), HE_oil_EMV.max(), HE_oil_EMV.min(), HE_oil_EMV.std()], start=1):
    ws6.cell(row=21, column=index, value=value)
for index, value in enumerate(
        ['GPR', Hedge_oil_GPR_OW.mean(), Hedge_oil_GPR_OW.max(), Hedge_oil_GPR_OW.min(), Hedge_oil_GPR_OW.std(),
         HR_oil_GPR.mean(), HR_oil_GPR.max(), HR_oil_GPR.min(), HR_oil_GPR.std(),
         HE_oil_GPR.mean(), HE_oil_GPR.max(), HE_oil_GPR.min(), HE_oil_GPR.std()], start=1):
    ws6.cell(row=22, column=index, value=value)
for index, value in enumerate(
        ['bs', Hedge_oil_bs_OW.mean(), Hedge_oil_bs_OW.max(), Hedge_oil_bs_OW.min(), Hedge_oil_bs_OW.std(),
         HR_oil_bs.mean(), HR_oil_bs.max(), HR_oil_bs.min(), HR_oil_bs.std(),
         HE_oil_bs.mean(), HE_oil_bs.max(), HE_oil_bs.min(), HE_oil_bs.std()], start=1):
    ws6.cell(row=23, column=index, value=value)

for index, value in enumerate(
        ['EPU', Hedge_btc_EPU_OW.mean(), Hedge_btc_EPU_OW.max(), Hedge_btc_EPU_OW.min(), Hedge_btc_EPU_OW.std(),
         HR_btc_EPU.mean(), HR_btc_EPU.max(), HR_btc_EPU.min(), HR_btc_EPU.std(),
         HE_btc_EPU.mean(), HE_btc_EPU.max(), HE_btc_EPU.min(), HE_btc_EPU.std()], start=1):
    ws6.cell(row=25, column=index, value=value)
for index, value in enumerate(
        ['CPU', Hedge_btc_CPU_OW.mean(), Hedge_btc_CPU_OW.max(), Hedge_btc_CPU_OW.min(), Hedge_btc_CPU_OW.std(),
         HR_btc_CPU.mean(), HR_btc_CPU.max(), HR_btc_CPU.min(), HR_btc_CPU.std(),
         HE_btc_CPU.mean(), HE_btc_CPU.max(), HE_btc_CPU.min(), HE_btc_CPU.std()], start=1):
    ws6.cell(row=26, column=index, value=value)
for index, value in enumerate(
        ['TPU', Hedge_btc_TPU_OW.mean(), Hedge_btc_TPU_OW.max(), Hedge_btc_TPU_OW.min(), Hedge_btc_TPU_OW.std(),
         HR_btc_TPU.mean(), HR_btc_TPU.max(), HR_btc_TPU.min(), HR_btc_TPU.std(),
         HE_btc_TPU.mean(), HE_btc_TPU.max(), HE_btc_TPU.min(), HE_btc_TPU.std()], start=1):
    ws6.cell(row=27, column=index, value=value)
for index, value in enumerate(
        ['EMV', Hedge_btc_EMV_OW.mean(), Hedge_btc_EMV_OW.max(), Hedge_btc_EMV_OW.min(), Hedge_btc_EMV_OW.std(),
         HR_btc_EMV.mean(), HR_btc_EMV.max(), HR_btc_EMV.min(), HR_btc_EMV.std(),
         HE_btc_EMV.mean(), HE_btc_EMV.max(), HE_btc_EMV.min(), HE_btc_EMV.std()], start=1):
    ws6.cell(row=28, column=index, value=value)
for index, value in enumerate(
        ['GPR', Hedge_btc_GPR_OW.mean(), Hedge_btc_GPR_OW.max(), Hedge_btc_GPR_OW.min(), Hedge_btc_GPR_OW.std(),
         HR_btc_GPR.mean(), HR_btc_GPR.max(), HR_btc_GPR.min(), HR_btc_GPR.std(),
         HE_btc_GPR.mean(), HE_btc_GPR.max(), HE_btc_GPR.min(), HE_btc_GPR.std()], start=1):
    ws6.cell(row=29, column=index, value=value)
for index, value in enumerate(
        ['bs', Hedge_btc_bs_OW.mean(), Hedge_btc_bs_OW.max(), Hedge_btc_bs_OW.min(), Hedge_btc_bs_OW.std(),
         HR_btc_bs.mean(), HR_btc_bs.max(), HR_btc_bs.min(), HR_btc_bs.std(),
         HE_btc_bs.mean(), HE_btc_bs.max(), HE_btc_bs.min(), HE_btc_bs.std()], start=1):
    ws6.cell(row=30, column=index, value=value)

wb6.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_Statistics_OWHRHE.xlsx')

##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************
##**********************************************************************************************************************

wb7 = Workbook()
ws7 = wb7.create_sheet("Statistics")

for index, value in enumerate(['Mean', 'Max.', 'Min.', 'Std. Dev.'] * 2, start=2):
    ws7.cell(row=2, column=index, value=value)
ws7['B1'] = 'Short-term'
ws7['F1'] = 'Long-term'

ws7['A3'] = 'Bond'
ws7['A9'] = 'Gold'
ws7['A15'] = 'Oil'
ws7['A21'] = 'Bitcoin'
for index, value in enumerate(['EPU', S_bond_EPU.mean(), S_bond_EPU.max(), S_bond_EPU.min(), S_bond_EPU.std(),
                               L_bond_EPU.mean(), L_bond_EPU.max(), L_bond_EPU.min(), L_bond_EPU.std()], start=1):
    ws7.cell(row=4, column=index, value=value)
for index, value in enumerate(['CPU', S_bond_CPU.mean(), S_bond_CPU.max(), S_bond_CPU.min(), S_bond_CPU.std(),
                               L_bond_CPU.mean(), L_bond_CPU.max(), L_bond_CPU.min(), L_bond_CPU.std()], start=1):
    ws7.cell(row=5, column=index, value=value)
for index, value in enumerate(['TPU', S_bond_TPU.mean(), S_bond_TPU.max(), S_bond_TPU.min(), S_bond_TPU.std(),
                               L_bond_TPU.mean(), L_bond_TPU.max(), L_bond_TPU.min(), L_bond_TPU.std()], start=1):
    ws7.cell(row=6, column=index, value=value)
for index, value in enumerate(['EMV', S_bond_EMV.mean(), S_bond_EMV.max(), S_bond_EMV.min(), S_bond_EMV.std(),
                               L_bond_EMV.mean(), L_bond_EMV.max(), L_bond_EMV.min(), L_bond_EMV.std()], start=1):
    ws7.cell(row=7, column=index, value=value)
for index, value in enumerate(['GPR', S_bond_GPR.mean(), S_bond_GPR.max(), S_bond_GPR.min(), S_bond_GPR.std(),
                               L_bond_GPR.mean(), L_bond_GPR.max(), L_bond_GPR.min(), L_bond_GPR.std()], start=1):
    ws7.cell(row=8, column=index, value=value)

for index, value in enumerate(['EPU', S_gold_EPU.mean(), S_gold_EPU.max(), S_gold_EPU.min(), S_gold_EPU.std(),
                               L_gold_EPU.mean(), L_gold_EPU.max(), L_gold_EPU.min(), L_gold_EPU.std()], start=1):
    ws7.cell(row=10, column=index, value=value)
for index, value in enumerate(['CPU', S_gold_CPU.mean(), S_gold_CPU.max(), S_gold_CPU.min(), S_gold_CPU.std(),
                               L_gold_CPU.mean(), L_gold_CPU.max(), L_gold_CPU.min(), L_gold_CPU.std()], start=1):
    ws7.cell(row=11, column=index, value=value)
for index, value in enumerate(['TPU', S_gold_TPU.mean(), S_gold_TPU.max(), S_gold_TPU.min(), S_gold_TPU.std(),
                               L_gold_TPU.mean(), L_gold_TPU.max(), L_gold_TPU.min(), L_gold_TPU.std()], start=1):
    ws7.cell(row=12, column=index, value=value)
for index, value in enumerate(['EMV', S_gold_EMV.mean(), S_gold_EMV.max(), S_gold_EMV.min(), S_gold_EMV.std(),
                               L_gold_EMV.mean(), L_gold_EMV.max(), L_gold_EMV.min(), L_gold_EMV.std()], start=1):
    ws7.cell(row=13, column=index, value=value)
for index, value in enumerate(['GPR', S_gold_GPR.mean(), S_gold_GPR.max(), S_gold_GPR.min(), S_gold_GPR.std(),
                               L_gold_GPR.mean(), L_gold_GPR.max(), L_gold_GPR.min(), L_gold_GPR.std()], start=1):
    ws7.cell(row=14, column=index, value=value)

for index, value in enumerate(['EPU', S_oil_EPU.mean(), S_oil_EPU.max(), S_oil_EPU.min(), S_oil_EPU.std(),
                               L_oil_EPU.mean(), L_oil_EPU.max(), L_oil_EPU.min(), L_oil_EPU.std()], start=1):
    ws7.cell(row=16, column=index, value=value)
for index, value in enumerate(['CPU', S_oil_CPU.mean(), S_oil_CPU.max(), S_oil_CPU.min(), S_oil_CPU.std(),
                               L_oil_CPU.mean(), L_oil_CPU.max(), L_oil_CPU.min(), L_oil_CPU.std()], start=1):
    ws7.cell(row=17, column=index, value=value)
for index, value in enumerate(['TPU', S_oil_TPU.mean(), S_oil_TPU.max(), S_oil_TPU.min(), S_oil_TPU.std(),
                               L_oil_TPU.mean(), L_oil_TPU.max(), L_oil_TPU.min(), L_oil_TPU.std()], start=1):
    ws7.cell(row=18, column=index, value=value)
for index, value in enumerate(['EMV', S_oil_EMV.mean(), S_oil_EMV.max(), S_oil_EMV.min(), S_oil_EMV.std(),
                               L_oil_EMV.mean(), L_oil_EMV.max(), L_oil_EMV.min(), L_oil_EMV.std()], start=1):
    ws7.cell(row=19, column=index, value=value)
for index, value in enumerate(['GPR', S_oil_GPR.mean(), S_oil_GPR.max(), S_oil_GPR.min(), S_oil_GPR.std(),
                               L_oil_GPR.mean(), L_oil_GPR.max(), L_oil_GPR.min(), L_oil_GPR.std()], start=1):
    ws7.cell(row=20, column=index, value=value)

for index, value in enumerate(['EPU', S_btc_EPU.mean(), S_btc_EPU.max(), S_btc_EPU.min(), S_btc_EPU.std(),
                               L_btc_EPU.mean(), L_btc_EPU.max(), L_btc_EPU.min(), L_btc_EPU.std()], start=1):
    ws7.cell(row=22, column=index, value=value)
for index, value in enumerate(['CPU', S_btc_CPU.mean(), S_btc_CPU.max(), S_btc_CPU.min(), S_btc_CPU.std(),
                               L_btc_CPU.mean(), L_btc_CPU.max(), L_btc_CPU.min(), L_btc_CPU.std()], start=1):
    ws7.cell(row=23, column=index, value=value)
for index, value in enumerate(['TPU', S_btc_TPU.mean(), S_btc_TPU.max(), S_btc_TPU.min(), S_btc_TPU.std(),
                               L_btc_TPU.mean(), L_btc_TPU.max(), L_btc_TPU.min(), L_btc_TPU.std()], start=1):
    ws7.cell(row=24, column=index, value=value)
for index, value in enumerate(['EMV', S_btc_EMV.mean(), S_btc_EMV.max(), S_btc_EMV.min(), S_btc_EMV.std(),
                               L_btc_EMV.mean(), L_btc_EMV.max(), L_btc_EMV.min(), L_btc_EMV.std()], start=1):
    ws7.cell(row=25, column=index, value=value)
for index, value in enumerate(['GPR', S_btc_GPR.mean(), S_btc_GPR.max(), S_btc_GPR.min(), S_btc_GPR.std(),
                               L_btc_GPR.mean(), L_btc_GPR.max(), L_btc_GPR.min(), L_btc_GPR.std()], start=1):
    ws7.cell(row=26, column=index, value=value)

wb7.save(r'C:\Users\徐浩然\Desktop\DCC-data-China\Hedge_Statistics_Cor.xlsx')
plt.show()


